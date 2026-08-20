"""Complete the isolated v7 backup/recovery fixture with files and boundary records."""

import hashlib
import io
import json
import uuid
import zipfile
from datetime import UTC, datetime, timedelta
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw
from pypdf import PdfWriter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.cli.recovery_v7_guard import RecoveryGuardError, require_recovery_environment
from app.cli.seed_capacity_demo import uid
from app.db.session import SessionLocal
from app.models import (
    AnalyticsSnapshot,
    Assignment,
    ClassStudent,
    GradeRelease,
    GradeReleaseItem,
    GradingBatch,
    KnowledgePoint,
    PaperPage,
    Question,
    QuestionKnowledgePoint,
    QuestionRubric,
    ReportJob,
    RubricItem,
    RubricVersion,
    SchoolClass,
    ScoreRevision,
    StoredFile,
    Student,
    Submission,
    SubmissionPage,
    SubmissionScoreSnapshot,
    TeacherReview,
    TeachingInsight,
    User,
    UserSession,
)
from app.results.services import create_analytics
from app.storage.dependencies import get_storage
from app.storage.minio import MinioStorage

SYNTHETIC_DOMAIN_SUFFIX = ".synthetic.invalid"


def fixture_bytes() -> dict[str, tuple[bytes, str, str]]:
    image = Image.new("RGB", (640, 360), "white")
    ImageDraw.Draw(image).text((30, 30), "AhaMark synthetic recovery page", fill="black")
    png = io.BytesIO()
    image.save(png, "PNG")
    pdf = io.BytesIO()
    writer = PdfWriter()
    writer.add_blank_page(width=595, height=842)
    writer.add_metadata({"/Title": "AhaMark 中文恢复验证"})
    writer.write(pdf)
    xlsx = io.BytesIO()
    workbook = Workbook()
    workbook.active.title = "恢复验证"
    workbook.active.append(["学号", "成绩"])
    workbook.active.append(["0001", 18])
    workbook.save(xlsx)
    archive = io.BytesIO()
    with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as bundle:
        bundle.writestr("student-0001.pdf", pdf.getvalue())
    return {
        "raw-paper": (png.getvalue(), "raw-paper.png", "image/png"),
        "submission-page": (png.getvalue(), "student-0001.png", "image/png"),
        "ocr-derived": (png.getvalue(), "ocr-processed.png", "image/png"),
        "gradebook": (
            xlsx.getvalue(),
            "成绩恢复验证.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        "chinese-pdf": (pdf.getvalue(), "中文恢复验证.pdf", "application/pdf"),
        "batch-zip": (archive.getvalue(), "批量报告.zip", "application/zip"),
        "other": (b'{"synthetic":true}', "manifest.json", "application/json"),
    }


def recovery_uid(run_id: str, name: str) -> uuid.UUID:
    return uuid.uuid5(uuid.NAMESPACE_URL, f"ahamark:recovery-v7:{run_id}:{name}")


def ensure_fixture_object(
    storage: MinioStorage,
    key: str,
    content: bytes,
    content_type: str,
    original_name: str,
) -> bytes:
    """Create once, or safely adopt equivalent current-run content after a DB rollback."""

    if storage.client.bucket_exists(storage.bucket):
        existing_keys = {
            item.object_name
            for item in storage.client.list_objects(storage.bucket, prefix=key, recursive=True)
            if item.object_name == key
        }
        if existing_keys:
            metadata = storage.stat(key)
            existing_content = storage.get(key).read()
            if metadata.content_type != content_type or not fixture_content_matches(
                original_name, existing_content, content
            ):
                raise RecoveryGuardError("existing fixture object does not match expected content")
            return existing_content
    storage.put(key, io.BytesIO(content), len(content), content_type)
    return content


def fixture_content_matches(name: str, actual: bytes, expected: bytes) -> bool:
    lower = name.lower()
    if lower.endswith(".xlsx"):

        def workbook_values(content: bytes) -> dict[str, list[list[object]]]:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
            return {
                sheet.title: [list(row) for row in sheet.iter_rows(values_only=True)]
                for sheet in workbook.worksheets
            }

        return workbook_values(actual) == workbook_values(expected)
    if lower.endswith(".zip"):

        def archive_values(content: bytes) -> dict[str, bytes]:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                return {name: archive.read(name) for name in sorted(archive.namelist())}

        return archive_values(actual) == archive_values(expected)
    return hashlib.sha256(actual).digest() == hashlib.sha256(expected).digest()


def assert_database_is_synthetic(db: Session, run_id: str) -> None:
    """Refuse a mixed or previously used database before adding fixture rows."""

    users = list(db.scalars(select(User)))
    if any(
        user.email is None or not user.email.endswith(SYNTHETIC_DOMAIN_SUFFIX) for user in users
    ):
        raise RecoveryGuardError("database contains a non-synthetic user")
    students = list(db.scalars(select(Student)))
    if any(not student.name.startswith("Synthetic ") for student in students):
        raise RecoveryGuardError("database contains a non-synthetic student")
    classes = list(db.scalars(select(SchoolClass)))
    if any(not school_class.name.startswith("Synthetic ") for school_class in classes):
        raise RecoveryGuardError("database contains a non-synthetic class")
    assignments = list(db.scalars(select(Assignment)))
    if any(not assignment.title.startswith("Synthetic ") for assignment in assignments):
        raise RecoveryGuardError("database contains a non-synthetic assignment")
    recovery_keys = list(
        db.scalars(
            select(StoredFile.storage_key).where(StoredFile.storage_key.like("recovery-v7/%"))
        )
    )
    expected_prefix = f"recovery-v7/{run_id}/"
    if any(not key.startswith(expected_prefix) for key in recovery_keys):
        raise RecoveryGuardError("database contains fixture objects from another recovery run")


def main() -> None:
    identity = require_recovery_environment()
    storage = get_storage()
    if not isinstance(storage, MinioStorage):
        raise RecoveryGuardError("recovery fixture requires isolated MinIO storage")
    teacher_id = uid("teacher-1")
    now = datetime.now(UTC)
    marker = f"recovery-v7-{identity.run_id}.synthetic.invalid"
    stored: dict[str, StoredFile] = {}
    with SessionLocal.begin() as db:
        assert_database_is_synthetic(db, identity.run_id)
        teacher = db.get(User, teacher_id)
        release = db.get(GradeRelease, uid("release-s1"))
        if teacher is None or release is None:
            raise RuntimeError("run seed_capacity_demo and seed_capacity_results first")
        if teacher.email is None or not teacher.email.endswith(SYNTHETIC_DOMAIN_SUFFIX):
            raise RecoveryGuardError("capacity teacher is not synthetic")
        if release.idempotency_key != "performance-capacity.synthetic.invalid:release:s1":
            raise RecoveryGuardError("capacity release marker is inconsistent")
        if db.scalar(select(UserSession).where(UserSession.user_id == teacher.id)) is None:
            db.add(
                UserSession(
                    user_id=teacher.id,
                    token_hash=hashlib.sha256(
                        f"synthetic-recovery-session:{identity.run_id}".encode()
                    ).hexdigest(),
                    csrf_hash=hashlib.sha256(
                        f"synthetic-recovery-csrf:{identity.run_id}".encode()
                    ).hexdigest(),
                    expires_at=now + timedelta(hours=1),
                    revoked_at=None,
                )
            )
        for kind, (content, name, content_type) in fixture_bytes().items():
            key = (
                f"{identity.object_prefix}{kind}/{recovery_uid(identity.run_id, f'object:{kind}')}"
            )
            record = db.scalar(select(StoredFile).where(StoredFile.storage_key == key))
            if record is None:
                stored_content = ensure_fixture_object(
                    storage,
                    key,
                    content,
                    content_type,
                    name,
                )
                record = StoredFile(
                    owner_id=teacher.id,
                    storage_key=key,
                    original_name=name,
                    content_type=content_type,
                    size=len(stored_content),
                    checksum=hashlib.sha256(stored_content).hexdigest(),
                    status="ready",
                )
                db.add(record)
                db.flush()
            stored[kind] = record
        paper_id = uid("paper-s1")
        if db.scalar(select(PaperPage).where(PaperPage.paper_version_id == paper_id)) is None:
            db.add(
                PaperPage(
                    paper_version_id=paper_id,
                    stored_file_id=stored["raw-paper"].id,
                    page_number=1,
                    source_page_number=1,
                    preview_storage_key=stored["ocr-derived"].storage_key,
                )
            )
        item = db.scalar(
            select(GradeReleaseItem)
            .where(GradeReleaseItem.grade_release_id == release.id)
            .order_by(GradeReleaseItem.created_at)
        )
        if item is None:
            raise RuntimeError("release has no included score")
        submission = db.get(Submission, item.submission_id)
        if submission is None:
            raise RuntimeError("release submission missing")
        if (
            db.scalar(select(SubmissionPage).where(SubmissionPage.submission_id == submission.id))
            is None
        ):
            db.add(
                SubmissionPage(
                    submission_id=submission.id,
                    stored_file_id=stored["submission-page"].id,
                    page_number=1,
                    source_page_number=1,
                    processed_storage_key=stored["ocr-derived"].storage_key,
                )
            )
        complete = db.get(SubmissionScoreSnapshot, item.score_snapshot_id)
        if complete is None:
            raise RuntimeError("release snapshot missing")
        rubric = db.get(RubricVersion, complete.rubric_version_id)
        question = db.scalar(
            select(Question)
            .where(Question.paper_version_id == complete.paper_version_id)
            .order_by(Question.display_order)
        )
        knowledge_point = db.scalar(
            select(KnowledgePoint)
            .where(KnowledgePoint.owner_id == teacher.id)
            .order_by(KnowledgePoint.id)
        )
        if rubric is None or question is None or knowledge_point is None:
            raise RecoveryGuardError("capacity rubric, question, or knowledge point is missing")
        question_rubric = db.scalar(
            select(QuestionRubric).where(
                QuestionRubric.rubric_version_id == rubric.id,
                QuestionRubric.question_id == question.id,
            )
        )
        expected_question_rubric_id = recovery_uid(identity.run_id, "question-rubric")
        if question_rubric is None:
            question_rubric = QuestionRubric(
                id=expected_question_rubric_id,
                rubric_version_id=rubric.id,
                question_id=question.id,
                standard_answer="Synthetic recovery answer",
                scoring_notes=f"Synthetic recovery rubric {identity.run_id}",
            )
            db.add(question_rubric)
            db.flush()
        elif question_rubric.id != expected_question_rubric_id:
            raise RecoveryGuardError("question rubric belongs to a different recovery run")
        rubric_item_id = recovery_uid(identity.run_id, "rubric-item")
        rubric_item = db.get(RubricItem, rubric_item_id)
        if rubric_item is None:
            db.add(
                RubricItem(
                    id=rubric_item_id,
                    question_rubric_id=question_rubric.id,
                    display_order=1,
                    title=f"Synthetic recovery criterion {identity.run_id}",
                    description="Synthetic recovery-only rubric item",
                    points=Decimal("1"),
                    required=True,
                )
            )
        elif rubric_item.question_rubric_id != question_rubric.id:
            raise RecoveryGuardError("rubric item relationship is inconsistent")
        question_point = db.get(
            QuestionKnowledgePoint,
            {"question_id": question.id, "knowledge_point_id": knowledge_point.id},
        )
        if question_point is None:
            db.add(
                QuestionKnowledgePoint(
                    question_id=question.id,
                    knowledge_point_id=knowledge_point.id,
                )
            )
        incomplete = db.scalar(
            select(SubmissionScoreSnapshot).where(
                SubmissionScoreSnapshot.submission_id == submission.id,
                SubmissionScoreSnapshot.version == 2,
            )
        )
        if incomplete is None:
            db.add(
                SubmissionScoreSnapshot(
                    id=recovery_uid(identity.run_id, "incomplete-snapshot"),
                    submission_id=submission.id,
                    assignment_id=complete.assignment_id,
                    student_id=complete.student_id,
                    paper_version_id=complete.paper_version_id,
                    rubric_version_id=complete.rubric_version_id,
                    total_score=None,
                    max_score=complete.max_score,
                    status="incomplete",
                    generated_by=teacher.id,
                    generated_at=now,
                    version=2,
                    details=[{"recovery_run_id": identity.run_id, "synthetic": True}],
                )
            )
        elif (
            incomplete.status != "incomplete"
            or incomplete.total_score is not None
            or incomplete.id != recovery_uid(identity.run_id, "incomplete-snapshot")
        ):
            raise RecoveryGuardError("existing incomplete snapshot is inconsistent with this run")
        review = db.scalar(select(TeacherReview).where(TeacherReview.reviewer_id == teacher.id))
        if (
            review is not None
            and db.scalar(select(ScoreRevision).where(ScoreRevision.teacher_review_id == review.id))
            is None
        ):
            db.add(
                ScoreRevision(
                    teacher_review_id=review.id,
                    student_answer_id=review.student_answer_id,
                    actor_id=teacher.id,
                    previous_score=review.final_score,
                    new_score=review.final_score,
                    previous_feedback=review.final_feedback,
                    new_feedback=review.final_feedback,
                    reason=f"Synthetic recovery audit boundary {identity.run_id}",
                )
            )
        report = db.scalar(select(ReportJob).where(ReportJob.idempotency_key == f"{marker}:xlsx"))
        if report is None:
            report = ReportJob(
                owner_id=teacher.id,
                assignment_id=release.assignment_id,
                class_id=release.class_id,
                grade_release_id=release.id,
                report_type="gradebook_xlsx",
                status="completed",
                progress=100,
                stored_file_id=stored["gradebook"].id,
                idempotency_key=f"{marker}:xlsx",
                started_at=now,
                completed_at=now,
            )
            db.add(report)
        analytics = db.scalar(
            select(AnalyticsSnapshot).where(AnalyticsSnapshot.grade_release_id == release.id)
        )
        if analytics is None:
            analytics = create_analytics(db, release)
        if (
            db.scalar(
                select(TeachingInsight).where(TeachingInsight.analytics_snapshot_id == analytics.id)
            )
            is None
        ):
            db.add(
                TeachingInsight(
                    owner_id=teacher.id,
                    analytics_snapshot_id=analytics.id,
                    content={
                        "title": "Synthetic rule-based recovery insight",
                        "recovery_run_id": identity.run_id,
                    },
                    evidence=[{"metric": "participant_count", "value": 50}],
                )
            )
        school_class = db.get(SchoolClass, release.class_id)
        batch = db.scalar(
            select(GradingBatch).where(
                GradingBatch.assignment_id == release.assignment_id,
                GradingBatch.class_id == release.class_id,
            )
        )
        if school_class is None or batch is None:
            raise RecoveryGuardError("capacity class or grading batch is missing")
        for kind in ("missing", "unfinished"):
            student_id = recovery_uid(identity.run_id, f"{kind}-student")
            student = db.get(Student, student_id)
            if student is None:
                compact_run_id = identity.run_id.replace("-", "").upper()
                student = Student(
                    id=student_id,
                    owner_id=teacher.id,
                    student_number=f"RV7{compact_run_id}{kind[:1].upper()}",
                    name=f"Synthetic Recovery {identity.run_id} {kind.title()} Student",
                )
                db.add(student)
                db.flush()
                db.add(
                    ClassStudent(
                        id=recovery_uid(identity.run_id, f"{kind}-membership"),
                        class_id=school_class.id,
                        student_id=student.id,
                    )
                )
            if kind == "unfinished":
                unfinished_id = recovery_uid(identity.run_id, "unfinished-submission")
                if db.get(Submission, unfinished_id) is None:
                    db.add(
                        Submission(
                            id=unfinished_id,
                            owner_id=teacher.id,
                            grading_batch_id=batch.id,
                            assignment_id=release.assignment_id,
                            class_id=release.class_id,
                            student_id=student.id,
                            status="uploaded",
                        )
                    )
    print(
        json.dumps(
            {
                "marker": marker,
                "run_id": identity.run_id,
                "objects": len(stored),
                "release_id": str(release.id),
            }
        )
    )


if __name__ == "__main__":
    main()
