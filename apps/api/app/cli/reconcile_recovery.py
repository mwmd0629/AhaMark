"""Read-only reconciliation plus fail-closed copying for isolated v7 recovery runs."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import time
import zipfile
from collections.abc import Iterable
from datetime import timedelta
from urllib.error import HTTPError
from urllib.request import urlopen

from minio import Minio
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader
from sqlalchemy import func, inspect, select, text
from sqlalchemy.orm import Session

from app.cli.recovery_v7_guard import RecoveryGuardError, require_recovery_environment
from app.core.config import get_settings
from app.db.session import SessionLocal
from app.models import (
    AnalyticsSnapshot,
    GradeRelease,
    GradeReleaseItem,
    PageProcessingResult,
    PaperPage,
    ReportJob,
    StoredFile,
    SubmissionPage,
    SubmissionScoreSnapshot,
    TeachingInsight,
)
from app.storage.dependencies import get_storage
from app.storage.minio import MinioStorage

TABLES = (
    "users",
    "user_sessions",
    "classes",
    "class_students",
    "students",
    "assignments",
    "paper_versions",
    "paper_pages",
    "questions",
    "reference_answer_versions",
    "structured_rubric_versions",
    "rubric_criteria",
    "structured_rubric_sets",
    "structured_rubric_set_items",
    "knowledge_points",
    "question_knowledge_points",
    "grading_batches",
    "submissions",
    "submission_pages",
    "student_answers",
    "teacher_reviews",
    "score_revisions",
    "submission_score_snapshots",
    "grade_releases",
    "grade_release_items",
    "report_jobs",
    "stored_files",
    "analytics_snapshots",
    "teaching_insights",
)
TRANSIENT_COLUMNS = frozenset({"created_at", "updated_at", "last_login_at"})


def stable_digest(value: object) -> str:
    canonical = json.dumps(value, default=str, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(canonical.encode()).hexdigest()


def stable_table_rows(db: Session, table_name: str) -> list[list[object]]:
    """Return deterministically ordered stable fields for a fixed, trusted table name."""

    inspector = inspect(db.get_bind())
    columns = [
        str(column["name"])
        for column in inspector.get_columns(table_name)
        if column["name"] not in TRANSIENT_COLUMNS
    ]
    primary_key = inspector.get_pk_constraint(table_name)
    primary_keys = [str(name) for name in primary_key["constrained_columns"]]
    order_columns = primary_keys or columns
    quoted_columns = ", ".join(f'"{column}"' for column in columns)
    order_by = ", ".join(f'"{column}"' for column in order_columns)
    rows = db.execute(text(f'SELECT {quoted_columns} FROM "{table_name}" ORDER BY {order_by}'))
    return [[row._mapping[column] for column in columns] for row in rows]


def stored_file_reference_violations(db: Session) -> dict[str, int]:
    """Discover every FK to stored_files and prove that all references resolve."""

    inspector = inspect(db.get_bind())
    violations: dict[str, int] = {}
    for table_name in inspector.get_table_names():
        for foreign_key in inspector.get_foreign_keys(table_name):
            if foreign_key.get("referred_table") != "stored_files":
                continue
            constrained = foreign_key.get("constrained_columns") or []
            referred = foreign_key.get("referred_columns") or []
            if len(constrained) != 1 or referred != ["id"]:
                raise RecoveryGuardError(f"unsupported StoredFile FK shape on {table_name}")
            column = str(constrained[0])
            count = db.execute(
                text(
                    f'SELECT count(*) FROM "{table_name}" source '
                    f'LEFT JOIN "stored_files" target ON source."{column}" = target.id '
                    f'WHERE source."{column}" IS NOT NULL AND target.id IS NULL'
                )
            ).scalar_one()
            violations[f"{table_name}.{column}"] = int(count)
    return violations


def stored_file_business_references(db: Session) -> dict[str, list[str]]:
    inspector = inspect(db.get_bind())
    references: dict[str, list[str]] = {}
    for table_name in inspector.get_table_names():
        primary_keys = inspector.get_pk_constraint(table_name)["constrained_columns"]
        for foreign_key in inspector.get_foreign_keys(table_name):
            if foreign_key.get("referred_table") != "stored_files":
                continue
            constrained = foreign_key.get("constrained_columns") or []
            if len(constrained) != 1:
                raise RecoveryGuardError(f"unsupported StoredFile FK shape on {table_name}")
            column = str(constrained[0])
            selected = [*primary_keys, column]
            projection = ", ".join(f'"{name}"' for name in selected)
            rows = db.execute(
                text(
                    f'SELECT {projection} FROM "{table_name}" '
                    f'WHERE "{column}" IS NOT NULL ORDER BY '
                    + ", ".join(f'"{name}"' for name in primary_keys)
                )
            )
            for row in rows:
                file_id = str(row._mapping[column])
                resource_id = ",".join(str(row._mapping[name]) for name in primary_keys)
                references.setdefault(file_id, []).append(f"{table_name}.{column}:{resource_id}")
    return {file_id: sorted(values) for file_id, values in sorted(references.items())}


def database_summary() -> dict[str, object]:
    identity = require_recovery_environment(
        database_roles=("source", "restored"),
        bucket_roles=("source",),
    )
    with SessionLocal() as db:
        counts = {
            table_name: int(db.execute(text(f'SELECT count(*) FROM "{table_name}"')).scalar_one())
            for table_name in TABLES
        }
        stable_rows = {table_name: stable_table_rows(db, table_name) for table_name in TABLES}
        snapshots = list(
            db.execute(
                select(
                    SubmissionScoreSnapshot.id,
                    SubmissionScoreSnapshot.submission_id,
                    SubmissionScoreSnapshot.version,
                    SubmissionScoreSnapshot.status,
                    SubmissionScoreSnapshot.total_score,
                ).order_by(SubmissionScoreSnapshot.id, SubmissionScoreSnapshot.version)
            )
        )
        release_items = list(
            db.execute(
                select(
                    GradeReleaseItem.grade_release_id,
                    GradeReleaseItem.student_id,
                    GradeReleaseItem.submission_id,
                    GradeReleaseItem.score_snapshot_id,
                ).order_by(GradeReleaseItem.grade_release_id, GradeReleaseItem.student_id)
            )
        )
        invalid_release_sources = db.scalar(
            select(func.count())
            .select_from(GradeReleaseItem)
            .join(
                SubmissionScoreSnapshot,
                SubmissionScoreSnapshot.id == GradeReleaseItem.score_snapshot_id,
            )
            .where(SubmissionScoreSnapshot.status != "complete")
        )
        invalid_reports = db.scalar(
            select(func.count())
            .select_from(ReportJob)
            .join(GradeRelease, GradeRelease.id == ReportJob.grade_release_id)
            .where(
                (ReportJob.assignment_id != GradeRelease.assignment_id)
                | (ReportJob.class_id != GradeRelease.class_id)
            )
        )
        invalid_analytics = db.scalar(
            select(func.count())
            .select_from(AnalyticsSnapshot)
            .join(GradeRelease, GradeRelease.id == AnalyticsSnapshot.grade_release_id)
            .where(
                (AnalyticsSnapshot.assignment_id != GradeRelease.assignment_id)
                | (AnalyticsSnapshot.class_id != GradeRelease.class_id)
                | (AnalyticsSnapshot.owner_id != GradeRelease.owner_id)
            )
        )
        invalid_insights = db.scalar(
            select(func.count())
            .select_from(TeachingInsight)
            .outerjoin(
                AnalyticsSnapshot,
                AnalyticsSnapshot.id == TeachingInsight.analytics_snapshot_id,
            )
            .where(AnalyticsSnapshot.id.is_(None))
        )
        duplicate_analytics = list(
            db.execute(
                select(AnalyticsSnapshot.grade_release_id, func.count())
                .where(AnalyticsSnapshot.status == "complete")
                .group_by(AnalyticsSnapshot.grade_release_id)
                .having(func.count() > 1)
            )
        )
        missing_students = int(
            db.execute(
                text(
                    "SELECT count(*) FROM class_students cs "
                    "JOIN grade_releases gr ON gr.class_id = cs.class_id "
                    "LEFT JOIN grade_release_items gri "
                    "ON gri.grade_release_id = gr.id AND gri.student_id = cs.student_id "
                    "WHERE gr.status = 'released' AND gri.id IS NULL"
                )
            ).scalar_one()
        )
        unfinalized_with_complete = int(
            db.execute(
                text(
                    "SELECT count(*) FROM submissions s "
                    "JOIN submission_score_snapshots ss ON ss.submission_id = s.id "
                    "WHERE s.status != 'finalized' AND ss.status = 'complete'"
                )
            ).scalar_one()
        )
        incomplete_zero_scores = int(
            db.execute(
                text(
                    "SELECT count(*) FROM submission_score_snapshots "
                    "WHERE status != 'complete' AND total_score = 0"
                )
            ).scalar_one()
        )
        revision = db.execute(text("SELECT version_num FROM alembic_version")).scalar_one()
        reference_violations = stored_file_reference_violations(db)
        business_references = stored_file_business_references(db)
        return {
            "run_id": identity.run_id,
            "alembic_revision": revision,
            "counts": counts,
            "key_ids": {
                table_name: [str(row[0]) for row in stable_rows[table_name]]
                for table_name in TABLES
                if stable_rows[table_name]
            },
            "stable_hash": stable_digest(stable_rows),
            "complete_snapshots": sum(row.status == "complete" for row in snapshots),
            "incomplete_snapshots": sum(row.status != "complete" for row in snapshots),
            "missing_students_without_formal_grade": missing_students,
            "unfinalized_students_with_complete_snapshot": unfinalized_with_complete,
            "incomplete_students_scored_as_zero": incomplete_zero_scores,
            "invalid_release_sources": int(invalid_release_sources or 0),
            "invalid_report_release_links": int(invalid_reports or 0),
            "invalid_analytics_release_links": int(invalid_analytics or 0),
            "invalid_insight_analytics_links": int(invalid_insights or 0),
            "duplicate_complete_analytics": [
                [str(value) for value in row] for row in duplicate_analytics
            ],
            "stored_file_reference_violations": reference_violations,
            "stored_file_business_references": business_references,
            "release_snapshot_bindings": [[str(value) for value in row] for row in release_items],
        }


def validate_structure(name: str, content: bytes) -> str:
    lower = name.lower()
    if lower.endswith(".png"):
        Image.open(io.BytesIO(content)).verify()
        return "png"
    if lower.endswith(".pdf"):
        if len(PdfReader(io.BytesIO(content)).pages) < 1:
            raise RuntimeError("PDF has no pages")
        return "pdf"
    if lower.endswith(".xlsx"):
        if not load_workbook(io.BytesIO(content), read_only=True).sheetnames:
            raise RuntimeError("XLSX has no worksheets")
        return "xlsx"
    if lower.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            if archive.testzip() is not None or not archive.namelist():
                raise RuntimeError("ZIP structure invalid")
        return "zip"
    if lower.endswith(".json"):
        json.loads(content)
        return "json"
    return "other"


def classify_orphans(
    *,
    objects: set[str],
    database_keys: set[str],
    legitimate_derived: set[str],
    known_historical: set[str],
    current_prefix: str,
) -> dict[str, list[str]]:
    missing_database = objects - database_keys
    unreferenced = objects - database_keys - legitimate_derived
    historical = unreferenced & known_historical
    unresolved = unreferenced - historical
    current_unknown = {key for key in unresolved if key.startswith(current_prefix)}
    unable = unresolved - current_unknown
    return {
        "database_missing_object": sorted(database_keys - objects),
        "object_missing_database": sorted(missing_database),
        "known_historical_orphans": sorted(historical),
        "legitimate_derived_objects": sorted(objects & legitimate_derived),
        "current_run_unknown_orphans": sorted(current_unknown),
        "unable_to_classify": sorted(unable),
    }


def ensure_no_target_objects(target_objects: Iterable[str]) -> None:
    existing = sorted(set(target_objects))
    if existing:
        raise RecoveryGuardError(
            f"recovery target contains {len(existing)} object(s); copying would overwrite"
        )


def derived_storage_references(db: Session) -> dict[str, list[str]]:
    references: dict[str, list[str]] = {}
    for model, columns in (
        (PaperPage, (PaperPage.preview_storage_key, PaperPage.thumbnail_storage_key)),
        (
            SubmissionPage,
            (
                SubmissionPage.rendered_storage_key,
                SubmissionPage.processed_storage_key,
                SubmissionPage.thumbnail_storage_key,
            ),
        ),
        (
            PageProcessingResult,
            (
                PageProcessingResult.original_storage_key,
                PageProcessingResult.rendered_storage_key,
                PageProcessingResult.processed_storage_key,
                PageProcessingResult.thumbnail_storage_key,
            ),
        ),
    ):
        for column in columns:
            rows = db.execute(select(model.id, column).where(column.is_not(None)))
            for resource_id, value in rows:
                if value is not None:
                    references.setdefault(value, []).append(
                        f"{model.__tablename__}.{column.key}:{resource_id}"
                    )
    return {key: sorted(values) for key, values in sorted(references.items())}


def read_object(client: Minio, bucket: str, key: str) -> bytes:
    response = client.get_object(bucket, key)
    try:
        return response.read()
    finally:
        response.close()
        response.release_conn()


def fetch_status(url: str) -> int:
    try:
        with urlopen(url, timeout=10) as response:  # noqa: S310 - test-only signed local URL
            response.read(1)
            return int(response.status)
    except HTTPError as exc:
        return int(exc.code)


def verify_signed_url(client: Minio, bucket: str, key: str, ttl_seconds: int) -> dict[str, object]:
    first_url = client.presigned_get_object(
        bucket,
        key,
        expires=timedelta(seconds=ttl_seconds),
    )
    initial_status = fetch_status(first_url)
    time.sleep(ttl_seconds + 1)
    expired_status = fetch_status(first_url)
    second_url = client.presigned_get_object(
        bucket,
        key,
        expires=timedelta(seconds=ttl_seconds),
    )
    renewed_status = fetch_status(second_url)
    old_after_renew_status = fetch_status(first_url)
    return {
        "ttl_seconds": ttl_seconds,
        "initial_status": initial_status,
        "expired_status": expired_status,
        "renewed_status": renewed_status,
        "old_after_renew_status": old_after_renew_status,
        "initial_read_succeeded": initial_status == 200,
        "old_url_expired": expired_status in {400, 403},
        "renewed_read_succeeded": renewed_status == 200,
        "old_url_remained_expired": old_after_renew_status in {400, 403},
        "url_query_recorded": False,
    }


def verify_signed_url_if_available(
    client: Minio,
    bucket: str,
    checked: list[dict[str, object]],
    ttl_seconds: int,
) -> dict[str, object]:
    if not checked:
        return {
            "skipped": True,
            "reason": "no_downloadable_objects",
            "url_query_recorded": False,
        }
    return verify_signed_url(client, bucket, str(checked[0]["object_key"]), ttl_seconds)


def object_reconcile(copy: bool) -> dict[str, object]:
    identity = require_recovery_environment()
    source = get_storage()
    if not isinstance(source, MinioStorage):
        raise RecoveryGuardError("source storage must be MinIO")
    target_bucket = os.environ.get("RECOVERY_TARGET_MINIO_BUCKET", "")
    if target_bucket != identity.restored_bucket:
        raise RecoveryGuardError("target MinIO bucket does not match this recovery run")
    target_endpoint = os.environ.get("RECOVERY_TARGET_MINIO_ENDPOINT", "")
    if not target_endpoint:
        raise RecoveryGuardError("RECOVERY_TARGET_MINIO_ENDPOINT is required")
    target = Minio(
        target_endpoint,
        access_key=os.environ["MINIO_ACCESS_KEY"],
        secret_key=os.environ["MINIO_SECRET_KEY"],
        secure=False,
    )
    public_endpoint = os.environ.get("RECOVERY_TARGET_MINIO_PUBLIC_ENDPOINT", target_endpoint)
    public_target = Minio(
        public_endpoint,
        access_key=os.environ["MINIO_ACCESS_KEY"],
        secret_key=os.environ["MINIO_SECRET_KEY"],
        secure=False,
    )
    with SessionLocal() as db:
        records = list(db.scalars(select(StoredFile).order_by(StoredFile.storage_key)))
        derived_references = derived_storage_references(db)
        file_references = stored_file_business_references(db)
    derived_keys = set(derived_references)
    record_by_key = {record.storage_key: record for record in records}
    source_objects = {
        item.object_name
        for item in source.client.list_objects(source.bucket, recursive=True)
        if item.object_name
    }
    foreign_source_objects = sorted(
        key for key in source_objects if not key.startswith(identity.object_prefix)
    )
    if foreign_source_objects:
        raise RecoveryGuardError("source bucket contains objects outside this recovery run")

    source_manifest: list[dict[str, object]] = []
    for key in sorted(source_objects):
        stat = source.client.stat_object(source.bucket, key)
        source_manifest.append(
            {
                "object_key": key,
                "size": int(stat.size or 0),
                "etag": stat.etag,
                "content_type": stat.content_type,
                "last_modified": stat.last_modified.isoformat() if stat.last_modified else None,
                "classification": (
                    "stored_file"
                    if key in record_by_key
                    else "legitimate_derived"
                    if key in derived_keys
                    else "unknown"
                ),
                "stored_file_id": (str(record_by_key[key].id) if key in record_by_key else None),
            }
        )
    if any(item["classification"] == "unknown" for item in source_manifest):
        raise RecoveryGuardError("source manifest contains an unclassified object")

    if copy:
        if target.bucket_exists(target_bucket):
            existing = {
                item.object_name
                for item in target.list_objects(target_bucket, recursive=True)
                if item.object_name
            }
            ensure_no_target_objects(existing)
            raise RecoveryGuardError("pre-existing target bucket lacks in-process creation proof")
        target.make_bucket(target_bucket)
        ensure_no_target_objects(
            item.object_name
            for item in target.list_objects(target_bucket, recursive=True)
            if item.object_name
        )
        for manifest_item in source_manifest:
            key = str(manifest_item["object_key"])
            content = read_object(source.client, source.bucket, key)
            target.put_object(
                target_bucket,
                key,
                io.BytesIO(content),
                len(content),
                content_type=str(manifest_item["content_type"] or "application/octet-stream"),
            )
    elif not target.bucket_exists(target_bucket):
        raise RecoveryGuardError("target bucket does not exist for read-only verification")

    target_objects = {
        item.object_name
        for item in target.list_objects(target_bucket, recursive=True)
        if item.object_name
    }
    checked: list[dict[str, object]] = []
    mismatches: list[dict[str, str]] = []
    for record in records:
        if record.storage_key not in target_objects:
            mismatches.append({"stored_file_id": str(record.id), "reason": "object_missing"})
            continue
        content = read_object(target, target_bucket, record.storage_key)
        stat = target.stat_object(target_bucket, record.storage_key)
        digest = hashlib.sha256(content).hexdigest()
        if len(content) != record.size or digest != record.checksum:
            mismatches.append({"stored_file_id": str(record.id), "reason": "size_or_checksum"})
        if stat.content_type != record.content_type:
            mismatches.append({"stored_file_id": str(record.id), "reason": "content_type"})
        checked.append(
            {
                "stored_file_id": str(record.id),
                "owner_id": str(record.owner_id),
                "object_key": record.storage_key,
                "size": len(content),
                "checksum": digest,
                "content_type": stat.content_type,
                "business_references": sorted(
                    file_references.get(str(record.id), [])
                    + derived_references.get(record.storage_key, [])
                ),
                "structure": validate_structure(record.original_name, content),
            }
        )
    historical = {
        value for value in os.environ.get("RECOVERY_KNOWN_HISTORICAL_KEYS", "").split(",") if value
    }
    orphan_report = classify_orphans(
        objects=target_objects,
        database_keys=set(record_by_key),
        legitimate_derived=derived_keys,
        known_historical=historical,
        current_prefix=identity.object_prefix,
    )
    settings = get_settings()
    signed_url = verify_signed_url_if_available(
        public_target,
        target_bucket,
        checked,
        settings.signed_url_expiry_seconds,
    )
    return {
        "run_id": identity.run_id,
        "source_bucket": source.bucket,
        "target_bucket": target_bucket,
        "database_records": len(records),
        "source_objects": len(source_objects),
        "target_objects": len(target_objects),
        "source_manifest": source_manifest,
        "checked": checked,
        "mismatches": mismatches,
        **orphan_report,
        "automatic_deletion": False,
        "signed_url_verification": signed_url,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("mode", choices=("database", "objects-copy", "objects-verify"))
    args = parser.parse_args()
    result = (
        database_summary()
        if args.mode == "database"
        else object_reconcile(args.mode == "objects-copy")
    )
    print(json.dumps(result, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
