import csv
import io
import re
import uuid
from datetime import UTC, datetime, timedelta
from typing import Annotated, Any, Literal

from app.api.actor import Actor
from app.core.config import get_settings
from app.db.session import get_db
from app.models import (
    ArchiveStatus,
    AuditLog,
    ClassStudent,
    ImportError,
    ImportJob,
    ImportRow,
    ImportRowStatus,
    ImportStatus,
    MembershipStatus,
    SchoolClass,
    Student,
    StudentGroup,
    StudentGroupMember,
    now_utc,
)
from app.security.files import UnsafeFile, inspect_xlsx_archive, safe_filename
from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
from pydantic import BaseModel, EmailStr, Field
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

router = APIRouter(prefix="/api", tags=["classes"])
Db = Annotated[Session, Depends(get_db)]


class ApiProblem(Exception):
    def __init__(self, status: int, code: str, message: str, details: dict[str, Any] | None = None):
        self.status, self.code, self.message, self.details = status, code, message, details or {}


class ClassInput(BaseModel):
    name: str = Field(min_length=1, max_length=120)
    grade: str | None = Field(default=None, max_length=40)
    subject: str | None = Field(default=None, max_length=40)
    academic_year: str | None = Field(default=None, max_length=20)
    semester: str | None = Field(default=None, max_length=30)
    description: str | None = Field(default=None, max_length=1000)


class ClassPatch(BaseModel):
    name: str | None = Field(default=None, min_length=1, max_length=120)
    grade: str | None = Field(default=None, max_length=40)
    subject: str | None = Field(default=None, max_length=40)
    academic_year: str | None = Field(default=None, max_length=20)
    semester: str | None = Field(default=None, max_length=30)
    description: str | None = Field(default=None, max_length=1000)


class StudentInput(BaseModel):
    name: str = Field(min_length=1, max_length=120)
    student_number: str = Field(min_length=1, max_length=64)
    gender: str | None = Field(default=None, max_length=20)
    email: EmailStr | None = None
    phone: str | None = Field(default=None, max_length=40)


class StudentPatch(BaseModel):
    name: str | None = Field(default=None, min_length=1, max_length=120)
    student_number: str | None = Field(default=None, min_length=1, max_length=64)
    gender: str | None = Field(default=None, max_length=20)
    email: EmailStr | None = None
    phone: str | None = Field(default=None, max_length=40)
    status: ArchiveStatus | None = None


class GroupInput(BaseModel):
    name: str = Field(min_length=1, max_length=80)
    description: str | None = Field(default=None, max_length=255)


class MemberInput(BaseModel):
    student_ids: list[uuid.UUID]


def audit(
    db: Session,
    actor_id: uuid.UUID,
    action: str,
    kind: str,
    rid: uuid.UUID,
    summary: dict[str, Any] | None = None,
) -> None:
    db.add(
        AuditLog(
            actor_id=actor_id,
            action=action,
            resource_type=kind,
            resource_id=str(rid),
            metadata_=summary or {},
        )
    )


def owned_class(db: Session, actor_id: uuid.UUID, class_id: uuid.UUID) -> SchoolClass:
    item = db.scalar(
        select(SchoolClass).where(SchoolClass.id == class_id, SchoolClass.owner_id == actor_id)
    )
    if item is None:
        raise ApiProblem(404, "CLASS_NOT_FOUND", "班级不存在")
    return item


def class_json(db: Session, item: SchoolClass) -> dict[str, Any]:
    active = (
        db.scalar(
            select(func.count())
            .select_from(ClassStudent)
            .where(ClassStudent.class_id == item.id, ClassStudent.status == MembershipStatus.active)
        )
        or 0
    )
    total = (
        db.scalar(
            select(func.count()).select_from(ClassStudent).where(ClassStudent.class_id == item.id)
        )
        or 0
    )
    groups = (
        db.scalar(
            select(func.count()).select_from(StudentGroup).where(StudentGroup.class_id == item.id)
        )
        or 0
    )
    return {
        "id": str(item.id),
        "name": item.name,
        "grade": item.grade,
        "subject": item.subject,
        "academic_year": item.academic_year,
        "semester": item.semester,
        "description": item.description,
        "status": item.status,
        "student_count": total,
        "active_student_count": active,
        "group_count": groups,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


@router.get("/classes")
def list_classes(
    db: Db,
    actor: Actor,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: str = "",
    status: ArchiveStatus | None = None,
    grade: str | None = None,
    subject: str | None = None,
    sort: Literal["updated_desc", "name_asc"] = "updated_desc",
) -> dict[str, Any]:
    filters: list[Any] = [SchoolClass.owner_id == actor.id]
    if search.strip():
        filters.append(SchoolClass.name.ilike(f"%{search.strip()}%"))
    if status:
        filters.append(SchoolClass.status == status)
    if grade:
        filters.append(SchoolClass.grade == grade)
    if subject:
        filters.append(SchoolClass.subject == subject)
    total = db.scalar(select(func.count()).select_from(SchoolClass).where(*filters)) or 0
    order = SchoolClass.name.asc() if sort == "name_asc" else SchoolClass.updated_at.desc()
    items = db.scalars(
        select(SchoolClass)
        .where(*filters)
        .order_by(order)
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    return {
        "items": [class_json(db, x) for x in items],
        "page": page,
        "page_size": page_size,
        "total": total,
        "pages": (total + page_size - 1) // page_size,
    }


@router.post("/classes", status_code=201)
def create_class(data: ClassInput, db: Db, actor: Actor) -> dict[str, Any]:
    name = data.name.strip()
    if db.scalar(
        select(SchoolClass.id).where(SchoolClass.owner_id == actor.id, SchoolClass.name == name)
    ):
        raise ApiProblem(409, "CLASS_NAME_CONFLICT", "当前教师已有同名班级", {"field": "name"})
    item = SchoolClass(owner_id=actor.id, **(data.model_dump() | {"name": name}))
    db.add(item)
    db.flush()
    audit(db, actor.id, "class.create", "class", item.id, {"name": name})
    db.commit()
    db.refresh(item)
    return class_json(db, item)


@router.get("/classes/{class_id}")
def get_class(class_id: uuid.UUID, db: Db, actor: Actor) -> dict[str, Any]:
    return class_json(db, owned_class(db, actor.id, class_id))


@router.patch("/classes/{class_id}")
def edit_class(class_id: uuid.UUID, data: ClassPatch, db: Db, actor: Actor) -> dict[str, Any]:
    item = owned_class(db, actor.id, class_id)
    changes = data.model_dump(exclude_unset=True)
    if "name" in changes:
        changes["name"] = changes["name"].strip()
        if db.scalar(
            select(SchoolClass.id).where(
                SchoolClass.owner_id == actor.id,
                SchoolClass.name == changes["name"],
                SchoolClass.id != item.id,
            )
        ):
            raise ApiProblem(409, "CLASS_NAME_CONFLICT", "当前教师已有同名班级", {"field": "name"})
    for key, value in changes.items():
        setattr(item, key, value)
    audit(db, actor.id, "class.update", "class", item.id, {"fields": sorted(changes)})
    db.commit()
    db.refresh(item)
    return class_json(db, item)


@router.post("/classes/{class_id}/archive")
def archive_class(class_id: uuid.UUID, db: Db, actor: Actor) -> dict[str, Any]:
    item = owned_class(db, actor.id, class_id)
    if item.status != ArchiveStatus.archived:
        item.status, item.archived_at = ArchiveStatus.archived, now_utc()
        audit(db, actor.id, "class.archive", "class", item.id)
        db.commit()
    return class_json(db, item)


@router.post("/classes/{class_id}/restore")
def restore_class(class_id: uuid.UUID, db: Db, actor: Actor) -> dict[str, Any]:
    item = owned_class(db, actor.id, class_id)
    if item.status != ArchiveStatus.active:
        item.status, item.archived_at = ArchiveStatus.active, None
        audit(db, actor.id, "class.restore", "class", item.id)
        db.commit()
    return class_json(db, item)


def student_json(
    db: Session, student: Student, membership: ClassStudent | None = None
) -> dict[str, Any]:
    groups = db.execute(
        select(StudentGroup.id, StudentGroup.name)
        .join(StudentGroupMember, StudentGroupMember.group_id == StudentGroup.id)
        .where(
            StudentGroupMember.student_id == student.id,
            *([StudentGroup.class_id == membership.class_id] if membership else []),
        )
    ).all()
    return {
        "id": str(student.id),
        "name": student.name,
        "student_number": student.student_number,
        "gender": student.gender,
        "email": student.email,
        "phone": student.phone,
        "status": student.status,
        "membership_status": membership.status if membership else None,
        "joined_at": membership.joined_at if membership else None,
        "groups": [{"id": str(g.id), "name": g.name} for g in groups],
        "assignment_history": [],
    }


@router.get("/classes/{class_id}/students")
def list_students(
    class_id: uuid.UUID,
    db: Db,
    actor: Actor,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    search: str = "",
    status: MembershipStatus | None = None,
    group_id: uuid.UUID | None = None,
    sort: Literal["joined_desc", "name_asc", "number_asc"] = "joined_desc",
) -> dict[str, Any]:
    owned_class(db, actor.id, class_id)
    q = (
        select(Student, ClassStudent)
        .join(ClassStudent, ClassStudent.student_id == Student.id)
        .where(ClassStudent.class_id == class_id, Student.owner_id == actor.id)
    )
    if search.strip():
        q = q.where(
            or_(
                Student.name.ilike(f"%{search.strip()}%"),
                Student.student_number.ilike(f"%{search.strip()}%"),
            )
        )
    if status:
        q = q.where(ClassStudent.status == status)
    if group_id:
        q = q.join(StudentGroupMember, StudentGroupMember.student_id == Student.id).where(
            StudentGroupMember.group_id == group_id
        )
    count = db.scalar(select(func.count()).select_from(q.subquery())) or 0
    order = (
        Student.name.asc()
        if sort == "name_asc"
        else Student.student_number.asc()
        if sort == "number_asc"
        else ClassStudent.joined_at.desc()
    )
    rows = db.execute(q.order_by(order).offset((page - 1) * page_size).limit(page_size)).all()
    return {
        "items": [student_json(db, s, m) for s, m in rows],
        "page": page,
        "page_size": page_size,
        "total": count,
        "pages": (count + page_size - 1) // page_size,
    }


@router.post("/classes/{class_id}/students", status_code=201)
def add_student(class_id: uuid.UUID, data: StudentInput, db: Db, actor: Actor) -> dict[str, Any]:
    owned_class(db, actor.id, class_id)
    number = data.student_number.strip()
    name = data.name.strip()
    student = db.scalar(
        select(Student).where(Student.owner_id == actor.id, Student.student_number == number)
    )
    created = student is None
    if student is None:
        student = Student(
            owner_id=actor.id, **(data.model_dump() | {"name": name, "student_number": number})
        )
        db.add(student)
        db.flush()
        audit(db, actor.id, "student.create", "student", student.id, {"student_number": number})
    membership = db.scalar(
        select(ClassStudent).where(
            ClassStudent.class_id == class_id, ClassStudent.student_id == student.id
        )
    )
    if membership and membership.status == MembershipStatus.active:
        raise ApiProblem(409, "STUDENT_ALREADY_IN_CLASS", "学生已在当前班级")
    if membership:
        membership.status, membership.removed_at, membership.joined_at = (
            MembershipStatus.active,
            None,
            now_utc(),
        )
    else:
        membership = ClassStudent(class_id=class_id, student_id=student.id)
        db.add(membership)
    audit(
        db,
        actor.id,
        "student.join_class",
        "student",
        student.id,
        {"class_id": str(class_id), "created": created},
    )
    db.commit()
    db.refresh(student)
    db.refresh(membership)
    return student_json(db, student, membership)


def owned_student(db: Session, actor_id: uuid.UUID, student_id: uuid.UUID) -> Student:
    item = db.scalar(select(Student).where(Student.id == student_id, Student.owner_id == actor_id))
    if item is None:
        raise ApiProblem(404, "STUDENT_NOT_FOUND", "学生不存在")
    return item


@router.get("/students/{student_id}")
def get_student(student_id: uuid.UUID, db: Db, actor: Actor) -> dict[str, Any]:
    return student_json(db, owned_student(db, actor.id, student_id))


@router.patch("/students/{student_id}")
def edit_student(student_id: uuid.UUID, data: StudentPatch, db: Db, actor: Actor) -> dict[str, Any]:
    item = owned_student(db, actor.id, student_id)
    changes = data.model_dump(exclude_unset=True)
    if "student_number" in changes:
        changes["student_number"] = changes["student_number"].strip()
        if db.scalar(
            select(Student.id).where(
                Student.owner_id == actor.id,
                Student.student_number == changes["student_number"],
                Student.id != item.id,
            )
        ):
            raise ApiProblem(
                409, "STUDENT_NUMBER_CONFLICT", "学号已存在", {"field": "student_number"}
            )
    if "name" in changes:
        changes["name"] = changes["name"].strip()
    for key, value in changes.items():
        setattr(item, key, value)
    if changes.get("status") == ArchiveStatus.archived:
        item.archived_at = now_utc()
    if changes.get("status") == ArchiveStatus.active:
        item.archived_at = None
    audit(db, actor.id, "student.update", "student", item.id, {"fields": sorted(changes)})
    db.commit()
    db.refresh(item)
    return student_json(db, item)


@router.delete("/classes/{class_id}/students/{student_id}")
def remove_student(
    class_id: uuid.UUID, student_id: uuid.UUID, db: Db, actor: Actor
) -> dict[str, str]:
    owned_class(db, actor.id, class_id)
    owned_student(db, actor.id, student_id)
    membership = db.scalar(
        select(ClassStudent).where(
            ClassStudent.class_id == class_id, ClassStudent.student_id == student_id
        )
    )
    if not membership or membership.status != MembershipStatus.active:
        raise ApiProblem(404, "STUDENT_NOT_IN_CLASS", "学生不在当前班级")
    membership.status, membership.removed_at = MembershipStatus.removed, now_utc()
    group_ids = select(StudentGroup.id).where(StudentGroup.class_id == class_id)
    for member in db.scalars(
        select(StudentGroupMember).where(
            StudentGroupMember.student_id == student_id, StudentGroupMember.group_id.in_(group_ids)
        )
    ).all():
        db.delete(member)
    audit(db, actor.id, "student.remove_class", "student", student_id, {"class_id": str(class_id)})
    db.commit()
    return {"status": "removed"}


@router.get("/classes/{class_id}/groups")
def groups(class_id: uuid.UUID, db: Db, actor: Actor) -> list[dict[str, Any]]:
    owned_class(db, actor.id, class_id)
    items = db.scalars(
        select(StudentGroup).where(StudentGroup.class_id == class_id).order_by(StudentGroup.name)
    ).all()
    return [
        {
            "id": str(x.id),
            "name": x.name,
            "description": x.description,
            "member_count": db.scalar(
                select(func.count())
                .select_from(StudentGroupMember)
                .where(StudentGroupMember.group_id == x.id)
            )
            or 0,
        }
        for x in items
    ]


@router.post("/classes/{class_id}/groups", status_code=201)
def create_group(class_id: uuid.UUID, data: GroupInput, db: Db, actor: Actor) -> dict[str, Any]:
    owned_class(db, actor.id, class_id)
    name = data.name.strip()
    if db.scalar(
        select(StudentGroup.id).where(StudentGroup.class_id == class_id, StudentGroup.name == name)
    ):
        raise ApiProblem(409, "GROUP_NAME_CONFLICT", "分组名称已存在", {"field": "name"})
    item = StudentGroup(class_id=class_id, name=name, description=data.description)
    db.add(item)
    db.flush()
    audit(db, actor.id, "group.create", "group", item.id, {"class_id": str(class_id)})
    db.commit()
    db.refresh(item)
    return {
        "id": str(item.id),
        "name": item.name,
        "description": item.description,
        "member_count": 0,
    }


def owned_group(db: Session, actor_id: uuid.UUID, group_id: uuid.UUID) -> StudentGroup:
    item = db.scalar(
        select(StudentGroup)
        .join(SchoolClass, SchoolClass.id == StudentGroup.class_id)
        .where(StudentGroup.id == group_id, SchoolClass.owner_id == actor_id)
    )
    if item is None:
        raise ApiProblem(404, "GROUP_NOT_FOUND", "分组不存在")
    return item


@router.patch("/groups/{group_id}")
def edit_group(group_id: uuid.UUID, data: GroupInput, db: Db, actor: Actor) -> dict[str, Any]:
    item = owned_group(db, actor.id, group_id)
    item.name, item.description = data.name.strip(), data.description
    audit(db, actor.id, "group.update", "group", item.id)
    db.commit()
    return {"id": str(item.id), "name": item.name, "description": item.description}


@router.delete("/groups/{group_id}", status_code=204)
def delete_group(group_id: uuid.UUID, db: Db, actor: Actor) -> None:
    item = owned_group(db, actor.id, group_id)
    audit(db, actor.id, "group.delete", "group", item.id)
    db.delete(item)
    db.commit()


@router.put("/groups/{group_id}/members")
def set_members(group_id: uuid.UUID, data: MemberInput, db: Db, actor: Actor) -> dict[str, int]:
    group = owned_group(db, actor.id, group_id)
    valid_ids = set(
        db.scalars(
            select(ClassStudent.student_id)
            .join(Student, Student.id == ClassStudent.student_id)
            .where(
                ClassStudent.class_id == group.class_id,
                ClassStudent.status == MembershipStatus.active,
                Student.owner_id == actor.id,
                ClassStudent.student_id.in_(data.student_ids),
            )
        ).all()
    )
    if valid_ids != set(data.student_ids):
        raise ApiProblem(409, "STUDENT_NOT_IN_CLASS", "只能选择当前班级中的学生")
    existing = db.scalars(
        select(StudentGroupMember).where(StudentGroupMember.group_id == group_id)
    ).all()
    for x in existing:
        db.delete(x)
    for sid in valid_ids:
        db.add(StudentGroupMember(group_id=group_id, student_id=sid))
    audit(db, actor.id, "group.members_set", "group", group.id, {"member_count": len(valid_ids)})
    db.commit()
    return {"member_count": len(valid_ids)}


HEADERS = {
    "姓名": "name",
    "学生姓名": "name",
    "学号": "student_number",
    "学生学号": "student_number",
    "分组": "group",
    "性别": "gender",
    "邮箱": "email",
    "联系方式": "phone",
    "联系电话": "phone",
}


@router.get("/import-template")
def import_template(format: Literal["xlsx", "csv"] = "xlsx") -> Response:
    headers = ["姓名", "学号", "分组", "性别", "邮箱", "联系方式"]
    if format == "csv":
        stream = io.StringIO()
        csv.writer(stream).writerow(headers)
        return Response(
            "\ufeff" + stream.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="student-import-template.csv"'},
        )
    book = Workbook()
    sheet = book.active
    sheet.title = "学生名单"
    sheet.append(headers)
    sheet.column_dimensions["B"].width = 18
    for cell in sheet["B"]:
        cell.number_format = numbers.FORMAT_TEXT
    output = io.BytesIO()
    book.save(output)
    return Response(
        output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="student-import-template.xlsx"'},
    )


def parse_upload(name: str, content: bytes) -> list[list[str]]:
    try:
        clean_name = safe_filename(name)
    except UnsafeFile as exc:
        raise ApiProblem(422, exc.code, exc.message) from exc
    suffix = clean_name.lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ApiProblem(422, "IMPORT_FILE_INVALID", "CSV 必须使用 UTF-8 编码") from exc
        return [[str(v).strip() for v in row] for row in csv.reader(io.StringIO(text))]
    if suffix == "xlsx":
        try:
            inspect_xlsx_archive(content)
            formula_sheet = load_workbook(
                io.BytesIO(content), read_only=True, data_only=False
            ).active
            for row in formula_sheet.iter_rows():
                if any(cell.data_type == "f" for cell in row):
                    raise ApiProblem(422, "IMPORT_FORMULA_FORBIDDEN", "导入表不得包含公式")
            sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
            return [
                ["" if v is None else str(v).strip() for v in row]
                for row in sheet.iter_rows(values_only=True)
            ]
        except ApiProblem:
            raise
        except UnsafeFile as exc:
            raise ApiProblem(422, exc.code, exc.message) from exc
        except Exception as exc:
            raise ApiProblem(422, "IMPORT_FILE_INVALID", "无法解析 XLSX 文件") from exc
    raise ApiProblem(415, "IMPORT_FILE_INVALID", "仅支持 .xlsx 和 .csv 文件")


@router.post("/classes/{class_id}/imports", status_code=201)
async def preview_import(
    class_id: uuid.UUID, db: Db, actor: Actor, file: Annotated[UploadFile, File()]
) -> dict[str, Any]:
    owned_class(db, actor.id, class_id)
    settings = get_settings()
    content = await file.read(settings.import_max_bytes + 1)
    if not content:
        raise ApiProblem(422, "IMPORT_FILE_INVALID", "文件为空")
    if len(content) > settings.import_max_bytes:
        raise ApiProblem(413, "IMPORT_TOO_LARGE", "文件超过 5MB 限制")
    rows = parse_upload(file.filename or "", content)
    while rows and not any(rows[-1]):
        rows.pop()
    if not rows:
        raise ApiProblem(422, "IMPORT_HEADERS_INVALID", "缺少表头")
    mapped = [HEADERS.get(x.strip()) for x in rows[0]]
    if "name" not in mapped or "student_number" not in mapped:
        raise ApiProblem(422, "IMPORT_HEADERS_INVALID", "表头必须包含姓名和学号")
    data_rows = [row for row in rows[1:] if any(v.strip() for v in row)]
    if len(data_rows) > settings.import_max_rows:
        raise ApiProblem(413, "IMPORT_TOO_LARGE", f"名单不能超过 {settings.import_max_rows} 行")
    job = ImportJob(
        owner_id=actor.id,
        class_id=class_id,
        original_name=(file.filename or "名单")[:255],
        file_type=(file.filename or "").rsplit(".", 1)[-1].lower(),
        status=ImportStatus.preview_ready,
        idempotency_key=uuid.uuid4().hex,
        expires_at=now_utc() + timedelta(hours=settings.import_expiry_hours),
    )
    db.add(job)
    db.flush()
    seen: set[str] = set()
    valid = invalid = duplicates = 0
    existing_groups = set(
        db.scalars(select(StudentGroup.name).where(StudentGroup.class_id == class_id)).all()
    )
    for row_no, cells in enumerate(data_rows, 2):
        normalized = {
            key: (cells[i].strip() if i < len(cells) else "") for i, key in enumerate(mapped) if key
        }
        dangerous = [
            key
            for key, value in normalized.items()
            if value.lstrip().startswith(("=", "+", "@"))
            or (
                value.lstrip().startswith("-")
                and len(value.lstrip()) > 1
                and value.lstrip()[1] in "=@+"
            )
        ]
        errors: list[tuple[str, str, str]] = []
        errors.extend(
            (key, "FORMULA_INJECTION", "字段不能以电子表格公式前缀开头") for key in dangerous
        )
        number = normalized.get("student_number", "")
        name = normalized.get("name", "")
        if not name:
            errors.append(("name", "NAME_REQUIRED", "姓名不能为空"))
        elif len(name) > 120:
            errors.append(("name", "NAME_TOO_LONG", "姓名不能超过 120 个字符"))
        if not number:
            errors.append(("student_number", "STUDENT_NUMBER_REQUIRED", "学号不能为空"))
        elif len(number) > 64:
            errors.append(("student_number", "STUDENT_NUMBER_TOO_LONG", "学号不能超过 64 个字符"))
        row_status = ImportRowStatus.valid
        if number and number in seen:
            errors.append(("student_number", "DUPLICATE_IN_FILE", "文件内学号重复"))
            row_status = ImportRowStatus.duplicate_in_file
            duplicates += 1
        seen.add(number)
        existing = (
            db.scalar(
                select(Student).where(
                    Student.owner_id == actor.id, Student.student_number == number
                )
            )
            if number
            else None
        )
        if existing:
            member = db.scalar(
                select(ClassStudent).where(
                    ClassStudent.class_id == class_id,
                    ClassStudent.student_id == existing.id,
                    ClassStudent.status == MembershipStatus.active,
                )
            )
            if member:
                errors.append(("student_number", "STUDENT_ALREADY_IN_CLASS", "该学生已在当前班级"))
                row_status = ImportRowStatus.duplicate_existing
                duplicates += 1
        email = normalized.get("email", "")
        if email and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
            errors.append(("email", "EMAIL_INVALID", "邮箱格式不正确"))
        group = normalized.get("group", "")
        if group and group not in existing_groups:
            errors.append(("group", "GROUP_NOT_FOUND", "分组不存在，请先创建分组"))
        if errors and row_status == ImportRowStatus.valid:
            row_status = ImportRowStatus.invalid
        record = ImportRow(
            import_job_id=job.id,
            row_number=row_no,
            raw_data={k: v for k, v in normalized.items() if k not in {"email", "phone"}},
            normalized_data=normalized,
            status=row_status,
        )
        db.add(record)
        for field, code, message in errors:
            db.add(
                ImportError(
                    import_job_id=job.id, row_number=row_no, field=field, code=code, message=message
                )
            )
        if errors:
            invalid += 1
        else:
            valid += 1
    job.total_rows, job.valid_rows, job.invalid_rows, job.duplicate_rows = (
        len(data_rows),
        valid,
        invalid,
        duplicates,
    )
    if valid == 0 and invalid:
        job.status = ImportStatus.validation_failed
    db.commit()
    return import_json(db, job)


def import_json(db: Session, job: ImportJob) -> dict[str, Any]:
    rows = db.scalars(
        select(ImportRow).where(ImportRow.import_job_id == job.id).order_by(ImportRow.row_number)
    ).all()
    errors = db.scalars(
        select(ImportError)
        .where(ImportError.import_job_id == job.id)
        .order_by(ImportError.row_number)
    ).all()
    by_row: dict[int, list[dict[str, str]]] = {}
    for e in errors:
        by_row.setdefault(e.row_number, []).append(
            {"field": e.field, "code": e.code, "message": e.message}
        )
    return {
        "id": str(job.id),
        "status": job.status,
        "original_name": job.original_name,
        "total_rows": job.total_rows,
        "valid_rows": job.valid_rows,
        "invalid_rows": job.invalid_rows,
        "duplicate_rows": job.duplicate_rows,
        "confirmed_rows": job.confirmed_rows,
        "result": job.result,
        "expires_at": job.expires_at,
        "rows": [
            {
                "row_number": r.row_number,
                "status": r.status,
                "data": r.normalized_data,
                "errors": by_row.get(r.row_number, []),
            }
            for r in rows
        ],
    }


def owned_import(db: Session, actor_id: uuid.UUID, import_id: uuid.UUID) -> ImportJob:
    job = db.scalar(
        select(ImportJob).where(ImportJob.id == import_id, ImportJob.owner_id == actor_id)
    )
    if job is None:
        raise ApiProblem(404, "IMPORT_NOT_FOUND", "导入任务不存在")
    return job


@router.get("/imports/{import_id}")
def get_import(import_id: uuid.UUID, db: Db, actor: Actor) -> dict[str, Any]:
    return import_json(db, owned_import(db, actor.id, import_id))


@router.post("/imports/{import_id}/confirm")
def confirm_import(import_id: uuid.UUID, db: Db, actor: Actor) -> dict[str, Any]:
    job = owned_import(db, actor.id, import_id)
    if job.status == ImportStatus.confirmed:
        return import_json(db, job)
    expiry = job.expires_at.replace(tzinfo=UTC) if job.expires_at.tzinfo is None else job.expires_at
    if expiry < datetime.now(UTC):
        raise ApiProblem(410, "IMPORT_EXPIRED", "导入预览已过期")
    if job.status not in {ImportStatus.preview_ready, ImportStatus.validation_failed}:
        raise ApiProblem(409, "IMPORT_NOT_READY", "导入任务尚未准备好")
    if job.invalid_rows or job.duplicate_rows:
        raise ApiProblem(
            409,
            "IMPORT_VALIDATION_FAILED",
            "导入预览包含错误或重复行，未写入任何学生；请修正文件后重新预览",
            {
                "invalid_rows": job.invalid_rows,
                "duplicate_rows": job.duplicate_rows,
            },
        )
    created = joined = assigned = 0
    for row in db.scalars(
        select(ImportRow).where(
            ImportRow.import_job_id == job.id, ImportRow.status == ImportRowStatus.valid
        )
    ).all():
        data = row.normalized_data
        student = db.scalar(
            select(Student).where(
                Student.owner_id == actor.id, Student.student_number == data["student_number"]
            )
        )
        if student is None:
            student = Student(
                owner_id=actor.id,
                name=data["name"],
                student_number=data["student_number"],
                gender=data.get("gender") or None,
                email=data.get("email") or None,
                phone=data.get("phone") or None,
            )
            db.add(student)
            db.flush()
            created += 1
        membership = db.scalar(
            select(ClassStudent).where(
                ClassStudent.class_id == job.class_id, ClassStudent.student_id == student.id
            )
        )
        if membership is None:
            membership = ClassStudent(class_id=job.class_id, student_id=student.id)
            db.add(membership)
            joined += 1
        elif membership.status == MembershipStatus.removed:
            membership.status, membership.removed_at = MembershipStatus.active, None
            joined += 1
        group_name = data.get("group")
        if group_name:
            group = db.scalar(
                select(StudentGroup).where(
                    StudentGroup.class_id == job.class_id, StudentGroup.name == group_name
                )
            )
            if group and not db.scalar(
                select(StudentGroupMember.id).where(
                    StudentGroupMember.group_id == group.id,
                    StudentGroupMember.student_id == student.id,
                )
            ):
                db.add(StudentGroupMember(group_id=group.id, student_id=student.id))
                assigned += 1
        row.status, row.student_id = ImportRowStatus.confirmed, student.id
    job.confirmed_rows = created + joined - min(created, joined)
    job.status = ImportStatus.confirmed
    job.confirmed_at = now_utc()
    job.result = {
        "created": created,
        "joined": joined,
        "skipped": job.invalid_rows,
        "failed": 0,
        "group_assignments": assigned,
    }
    audit(db, actor.id, "import.confirm", "import_job", job.id, job.result)
    db.commit()
    return import_json(db, job)
