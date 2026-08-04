from __future__ import annotations

import io
import statistics
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field, ValidationError, field_validator, model_validator
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table
from sqlalchemy import select, update
from sqlalchemy.orm import Session

from app.models import (
    AnalyticsSnapshot,
    Assignment,
    GradeRelease,
    GradeReleaseItem,
    KnowledgePoint,
    PaperVersion,
    Question,
    RubricVersion,
    StudentAnswer,
    Submission,
    SubmissionScoreSnapshot,
    TeacherReview,
    now_utc,
)


def serialize_grade_release_mutation(
    db: Session, owner_id: uuid.UUID, assignment_id: uuid.UUID
) -> bool:
    """Serialize release creation/publication on one assignment transaction.

    The self-update is intentional: PostgreSQL takes an assignment row lock and
    SQLite takes its write lock without changing the assignment fingerprint.
    Every release mutation must acquire this lock before narrower rows.
    """
    locked_id = db.scalar(
        update(Assignment)
        .where(Assignment.id == assignment_id, Assignment.owner_id == owner_id)
        .values(updated_at=Assignment.updated_at)
        .returning(Assignment.id)
    )
    return locked_id is not None


class SnapshotDetail(BaseModel):
    model_config = ConfigDict(extra="allow")
    question_id: uuid.UUID
    question_number: str
    question_type: str
    score: Decimal
    max_score: Decimal
    teacher_review_id: uuid.UUID
    final_error_type: str | None = None
    final_feedback: str | None = None
    knowledge_point_ids: list[uuid.UUID] = Field(default_factory=list)
    grading_method: str
    finalized_at: datetime

    @field_validator("max_score")
    @classmethod
    def positive_maximum(cls, value: Decimal) -> Decimal:
        if value <= 0:
            raise ValueError("max_score must be positive")
        return value

    @model_validator(mode="after")
    def legal_score(self) -> SnapshotDetail:
        if self.score < 0 or self.score > self.max_score:
            raise ValueError("score is outside 0..max_score")
        return self


class SnapshotPayload(BaseModel):
    schema_version: str = "1.0"
    submission_id: uuid.UUID
    assignment_id: uuid.UUID
    student_id: uuid.UUID
    paper_version_id: uuid.UUID
    rubric_version_id: uuid.UUID
    total_score: Decimal
    max_score: Decimal
    question_count: int
    details: list[SnapshotDetail]

    @model_validator(mode="after")
    def internally_consistent(self) -> SnapshotPayload:
        if self.question_count <= 0 or not self.details:
            raise ValueError("snapshot must contain at least one question")
        if self.max_score <= 0:
            raise ValueError("max_score must be positive")
        ids = [item.question_id for item in self.details]
        if len(ids) != len(set(ids)):
            raise ValueError("duplicate question_id")
        if self.question_count != len(self.details):
            raise ValueError("question_count mismatch")
        if self.total_score != sum((x.score for x in self.details), Decimal(0)):
            raise ValueError("total_score mismatch")
        if self.max_score != sum((x.max_score for x in self.details), Decimal(0)):
            raise ValueError("max_score mismatch")
        return self


@dataclass(frozen=True)
class ValidatedScore:
    snapshot: SubmissionScoreSnapshot
    submission: Submission
    payload: SnapshotPayload


class FinalScoreService:
    """The only supported final-score reader. It never queries GradingResult."""

    def __init__(self, db: Session, owner_id: uuid.UUID):
        self.db, self.owner_id = db, owner_id

    def latest(self, assignment_id: uuid.UUID, class_id: uuid.UUID) -> list[ValidatedScore]:
        submissions = self.db.scalars(
            select(Submission).where(
                Submission.owner_id == self.owner_id,
                Submission.assignment_id == assignment_id,
                Submission.class_id == class_id,
                Submission.status == "finalized",
            )
        ).all()
        latest_by_student: dict[uuid.UUID, ValidatedScore] = {}
        for submission in submissions:
            snapshots = self.db.scalars(
                select(SubmissionScoreSnapshot)
                .where(
                    SubmissionScoreSnapshot.submission_id == submission.id,
                    SubmissionScoreSnapshot.status == "complete",
                )
                .order_by(
                    SubmissionScoreSnapshot.version.desc(),
                    SubmissionScoreSnapshot.generated_at.desc(),
                    SubmissionScoreSnapshot.id.desc(),
                )
            ).all()
            score: ValidatedScore | None = None
            for snapshot in snapshots:
                try:
                    score = self.validate(snapshot, submission)
                    break
                except ValueError:
                    continue
            if score is None:
                continue
            student_id = score.payload.student_id
            current = latest_by_student.get(student_id)
            if current is None or (
                score.snapshot.generated_at,
                score.snapshot.version,
                str(score.snapshot.id),
            ) > (
                current.snapshot.generated_at,
                current.snapshot.version,
                str(current.snapshot.id),
            ):
                latest_by_student[student_id] = score
        return list(latest_by_student.values())

    def validate(self, snapshot: SubmissionScoreSnapshot, submission: Submission) -> ValidatedScore:
        if (
            submission.owner_id != self.owner_id
            or submission.status != "finalized"
            or submission.finalized_at is None
        ):
            raise ValueError("SUBMISSION_NOT_FINALIZED_OR_OWNED")
        if snapshot.status != "complete":
            raise ValueError("SNAPSHOT_NOT_COMPLETE")
        if snapshot.generated_by != self.owner_id:
            raise ValueError("SNAPSHOT_OWNER_MISMATCH")
        raw = {
            "schema_version": "1.0",
            "submission_id": snapshot.submission_id,
            "assignment_id": snapshot.assignment_id,
            "student_id": snapshot.student_id,
            "paper_version_id": snapshot.paper_version_id,
            "rubric_version_id": snapshot.rubric_version_id,
            "total_score": snapshot.total_score,
            "max_score": snapshot.max_score,
            "question_count": len(snapshot.details),
            "details": snapshot.details,
        }
        try:
            payload = SnapshotPayload.model_validate(raw)
        except (ValidationError, InvalidOperation) as exc:
            raise ValueError(f"SNAPSHOT_SCHEMA_INVALID: {exc}") from exc
        if (
            payload.submission_id != submission.id
            or payload.assignment_id != submission.assignment_id
            or payload.student_id != submission.student_id
        ):
            raise ValueError("SNAPSHOT_RELATION_MISMATCH")
        paper = self.db.get(PaperVersion, payload.paper_version_id)
        rubric = self.db.get(RubricVersion, payload.rubric_version_id)
        if (
            paper is None
            or paper.assignment_id != payload.assignment_id
            or rubric is None
            or rubric.assignment_id != payload.assignment_id
        ):
            raise ValueError("SNAPSHOT_VERSION_RELATION_MISMATCH")
        questions = {
            question.id: question
            for question in self.db.scalars(
                select(Question).where(
                    Question.id.in_([x.question_id for x in payload.details]),
                    Question.paper_version_id == payload.paper_version_id,
                )
            )
        }
        if set(questions) != {x.question_id for x in payload.details}:
            raise ValueError("SNAPSHOT_QUESTION_MISSING")
        reviews = {
            review.id: review
            for review in self.db.scalars(
                select(TeacherReview).where(
                    TeacherReview.id.in_([x.teacher_review_id for x in payload.details])
                )
            )
        }
        answers = {
            answer.id: answer
            for answer in self.db.scalars(
                select(StudentAnswer).where(
                    StudentAnswer.id.in_([x.student_answer_id for x in reviews.values()])
                )
            )
        }
        for detail in payload.details:
            question = questions[detail.question_id]
            review = reviews.get(detail.teacher_review_id)
            answer = answers.get(review.student_answer_id) if review else None
            if (
                question.question_number != detail.question_number
                or question.question_type != detail.question_type
                or question.max_score is None
                or Decimal(question.max_score) != detail.max_score
            ):
                raise ValueError("SNAPSHOT_QUESTION_METADATA_MISMATCH")
            if (
                review is None
                or review.confirmed_at is None
                or answer is None
                or answer.submission_id != submission.id
                or answer.question_id != detail.question_id
            ):
                raise ValueError("SNAPSHOT_REVIEW_RELATION_MISMATCH")
        knowledge_ids = {
            value for detail in payload.details for value in detail.knowledge_point_ids
        }
        if knowledge_ids:
            known = set(
                self.db.scalars(
                    select(KnowledgePoint.id).where(
                        KnowledgePoint.id.in_(knowledge_ids),
                        KnowledgePoint.owner_id == self.owner_id,
                    )
                )
            )
            if known != knowledge_ids:
                raise ValueError("SNAPSHOT_KNOWLEDGE_POINT_MISMATCH")
        return ValidatedScore(snapshot, submission, payload)

    def validate_released(
        self, snapshot: SubmissionScoreSnapshot, submission: Submission
    ) -> ValidatedScore:
        """Validate an immutable release source without consulting mutable workflow state."""
        if submission.owner_id != self.owner_id:
            raise ValueError("SUBMISSION_OWNER_MISMATCH")
        if snapshot.status != "complete":
            raise ValueError("SNAPSHOT_NOT_COMPLETE")
        if snapshot.generated_by != self.owner_id:
            raise ValueError("SNAPSHOT_OWNER_MISMATCH")
        raw = {
            "schema_version": "1.0",
            "submission_id": snapshot.submission_id,
            "assignment_id": snapshot.assignment_id,
            "student_id": snapshot.student_id,
            "paper_version_id": snapshot.paper_version_id,
            "rubric_version_id": snapshot.rubric_version_id,
            "total_score": snapshot.total_score,
            "max_score": snapshot.max_score,
            "question_count": len(snapshot.details),
            "details": snapshot.details,
        }
        try:
            payload = SnapshotPayload.model_validate(raw)
        except (ValidationError, InvalidOperation) as exc:
            raise ValueError(f"SNAPSHOT_SCHEMA_INVALID: {exc}") from exc
        if (
            payload.submission_id != submission.id
            or payload.assignment_id != submission.assignment_id
            or payload.student_id != submission.student_id
        ):
            raise ValueError("SNAPSHOT_RELATION_MISMATCH")
        paper = self.db.get(PaperVersion, payload.paper_version_id)
        rubric = self.db.get(RubricVersion, payload.rubric_version_id)
        if (
            paper is None
            or paper.assignment_id != payload.assignment_id
            or rubric is None
            or rubric.assignment_id != payload.assignment_id
        ):
            raise ValueError("SNAPSHOT_VERSION_RELATION_MISMATCH")
        question_ids = {detail.question_id for detail in payload.details}
        questions = set(
            self.db.scalars(
                select(Question.id).where(
                    Question.id.in_(question_ids),
                    Question.paper_version_id == payload.paper_version_id,
                )
            )
        )
        if questions != question_ids:
            raise ValueError("SNAPSHOT_QUESTION_MISSING")
        reviews = {
            review.id: review
            for review in self.db.scalars(
                select(TeacherReview).where(
                    TeacherReview.id.in_([x.teacher_review_id for x in payload.details])
                )
            )
        }
        answers = {
            answer.id: answer
            for answer in self.db.scalars(
                select(StudentAnswer).where(
                    StudentAnswer.id.in_([x.student_answer_id for x in reviews.values()])
                )
            )
        }
        for detail in payload.details:
            review = reviews.get(detail.teacher_review_id)
            answer = answers.get(review.student_answer_id) if review else None
            if (
                review is None
                or review.confirmed_at is None
                or answer is None
                or answer.submission_id != submission.id
                or answer.question_id != detail.question_id
            ):
                raise ValueError("SNAPSHOT_REVIEW_RELATION_MISMATCH")
        knowledge_ids = {
            value for detail in payload.details for value in detail.knowledge_point_ids
        }
        if knowledge_ids:
            known = set(
                self.db.scalars(
                    select(KnowledgePoint.id).where(
                        KnowledgePoint.id.in_(knowledge_ids),
                        KnowledgePoint.owner_id == self.owner_id,
                    )
                )
            )
            if known != knowledge_ids:
                raise ValueError("SNAPSHOT_KNOWLEDGE_POINT_MISMATCH")
        return ValidatedScore(snapshot, submission, payload)


def compute_metrics(
    scores: list[ValidatedScore],
    knowledge_point_names: dict[str, str] | None = None,
) -> dict[str, Any]:
    knowledge_point_names = knowledge_point_names or {}
    totals = [float(x.payload.total_score) for x in scores]
    ratios = [float(x.payload.total_score / x.payload.max_score) for x in scores]
    questions: dict[str, dict[str, Any]] = {}
    knowledge: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"score": 0.0, "max": 0.0, "questions": set(), "participants": set()}
    )
    errors: Counter[str] = Counter()
    for row in scores:
        for detail in row.payload.details:
            key = str(detail.question_id)
            item = questions.setdefault(
                key,
                {
                    "question_id": key,
                    "question_number": detail.question_number,
                    "question_type": detail.question_type,
                    "score": 0.0,
                    "max": 0.0,
                    "full": 0,
                    "zero": 0,
                    "participants": 0,
                },
            )
            item["score"] += float(detail.score)
            item["max"] += float(detail.max_score)
            item["participants"] += 1
            item["full"] += detail.score == detail.max_score
            item["zero"] += detail.score == 0
            if detail.final_error_type:
                errors[detail.final_error_type] += 1
            for knowledge_id in set(detail.knowledge_point_ids):
                kp = knowledge[str(knowledge_id)]
                kp["score"] += float(detail.score)
                kp["max"] += float(detail.max_score)
                kp["questions"].add(key)
                kp["participants"].add(str(row.payload.student_id))
    question_rows = []
    for item in questions.values():
        count = item["participants"]
        item.update(
            average_score=item["score"] / count,
            average_max_score=item["max"] / count,
            score_rate=item["score"] / item["max"] if item["max"] else None,
            full_rate=item["full"] / count,
            zero_rate=item["zero"] / count,
            correct_rate=item["full"] / count
            if item["question_type"]
            in {"single_choice", "multiple_choice", "true_false", "fill_blank"}
            else None,
        )
        question_rows.append(item)
    distribution = {"0-59": 0, "60-69": 0, "70-79": 0, "80-89": 0, "90-100": 0}
    layers = {"A": 0, "B": 0, "C": 0, "D": 0}
    for ratio in ratios:
        pct = ratio * 100
        distribution[
            "0-59"
            if pct < 60
            else "60-69"
            if pct < 70
            else "70-79"
            if pct < 80
            else "80-89"
            if pct < 90
            else "90-100"
        ] += 1
        layers[
            "A" if ratio >= 0.85 else "B" if ratio >= 0.70 else "C" if ratio >= 0.50 else "D"
        ] += 1
    return {
        "metric_version": "1.0",
        "participant_count": len(scores),
        "average_score": statistics.fmean(totals) if totals else None,
        "highest_score": max(totals) if totals else None,
        "lowest_score": min(totals) if totals else None,
        "median_score": statistics.median(totals) if totals else None,
        "average_score_rate": statistics.fmean(ratios) if ratios else None,
        "score_distribution": distribution,
        "student_layers": layers,
        "questions": sorted(question_rows, key=lambda x: x["question_number"]),
        "knowledge_points": [
            {
                "knowledge_point_id": key,
                "knowledge_point_name": knowledge_point_names.get(key, key),
                "mastery_rate": value["score"] / value["max"] if value["max"] else None,
                "question_ids": sorted(value["questions"]),
                "sample_count": len(value["participants"]),
            }
            for key, value in knowledge.items()
        ],
        "error_types": [{"code": code, "count": count} for code, count in errors.most_common()],
    }


def release_scores(db: Session, release_id: uuid.UUID) -> list[ValidatedScore]:
    release = db.get(GradeRelease, release_id)
    if release is None:
        raise ValueError("RELEASE_NOT_FOUND")
    if release.status != "released":
        raise ValueError("RELEASE_NOT_ACTIVE")
    if release.created_by != release.owner_id:
        raise ValueError("RELEASE_OWNER_MISMATCH")
    service = FinalScoreService(db, release.owner_id)
    rows: list[ValidatedScore] = []
    for item in db.scalars(
        select(GradeReleaseItem).where(
            GradeReleaseItem.grade_release_id == release.id, GradeReleaseItem.status == "included"
        )
    ):
        snapshot, submission = (
            db.get(SubmissionScoreSnapshot, item.score_snapshot_id),
            db.get(Submission, item.submission_id),
        )
        if snapshot is None or submission is None:
            raise ValueError("RELEASE_SOURCE_MISSING")
        score = service.validate_released(snapshot, submission)
        if (
            item.student_id != score.payload.student_id
            or item.submission_id != score.payload.submission_id
            or item.score_snapshot_id != score.snapshot.id
            or submission.assignment_id != release.assignment_id
            or submission.class_id != release.class_id
        ):
            raise ValueError("RELEASE_ITEM_RELATION_MISMATCH")
        rows.append(score)
    return rows


def gradebook_xlsx(db: Session, release: GradeRelease) -> bytes:
    scores = release_scores(db, release.id)
    metrics = compute_metrics(scores)
    questions = sorted({d.question_number for row in scores for d in row.payload.details})

    def safe(value: str) -> str:
        return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value

    wb = Workbook()
    ws = wb.active
    ws.title = "成绩总表"
    ws.append(
        [
            "学号",
            "姓名",
            "班级",
            "作业",
            *[safe(question) for question in questions],
            "总分",
            "满分",
            "得分率",
            "成绩状态",
            "最终确认时间",
        ]
    )
    for row in scores:
        from app.models import Assignment, SchoolClass, Student

        student, school_class, assignment = (
            db.get(Student, row.payload.student_id),
            db.get(SchoolClass, release.class_id),
            db.get(Assignment, release.assignment_id),
        )
        detail = {x.question_number: float(x.score) for x in row.payload.details}

        ws.append(
            [
                safe(student.student_number) if student else "",
                safe(student.name) if student else "",
                safe(school_class.name) if school_class else "",
                safe(assignment.title) if assignment else "",
                *[detail.get(q) for q in questions],
                float(row.payload.total_score),
                float(row.payload.max_score),
                float(row.payload.total_score / row.payload.max_score),
                "已发布",
                row.snapshot.generated_at.isoformat(),
            ]
        )
        ws.cell(ws.max_row, 1).number_format = "@"
    qws = wb.create_sheet("题目统计")
    qws.append(["题号", "题型", "参与人数", "平均得分率", "满分率", "零分率", "正确率"])
    for q in metrics["questions"]:
        qws.append(
            [
                safe(q["question_number"]),
                q["question_type"],
                q["participants"],
                q["score_rate"],
                q["full_rate"],
                q["zero_rate"],
                q["correct_rate"],
            ]
        )
    kws = wb.create_sheet("知识点统计")
    kws.append(["知识点ID", "关联题目", "参与人数", "掌握率", "样本说明"])
    for kp in metrics["knowledge_points"]:
        name = db.get(KnowledgePoint, uuid.UUID(kp["knowledge_point_id"]))
        kws.append(
            [
                safe(name.name) if name else kp["knowledge_point_id"],
                ",".join(kp["question_ids"]),
                kp["sample_count"],
                kp["mastery_rate"],
                "一题多知识点时完整计入每个知识点",
            ]
        )
    notes = wb.create_sheet("导出说明")
    notes.append(["发布版本", release.version])
    notes.append(["数据规则", "仅使用发布批次固定的 complete ScoreSnapshot"])
    notes.append(["缺失规则", "未完成不记为零分"])
    notes.append(["系统版本", "AhaMark 0.1.0 / metrics 1.0"])
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 18
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def student_report_pdf(
    db: Session, release: GradeRelease, student_id: uuid.UUID, font_path: Path
) -> bytes:
    from app.models import Assignment, SchoolClass, Student

    scores = [row for row in release_scores(db, release.id) if row.payload.student_id == student_id]
    if len(scores) != 1:
        raise ValueError("发布版本中不存在唯一学生成绩")
    row = scores[0]
    student = db.get(Student, student_id)
    school_class = db.get(SchoolClass, release.class_id)
    assignment = db.get(Assignment, release.assignment_id)
    if student is None or school_class is None or assignment is None:
        raise ValueError("报告引用数据不完整")
    font_name = "NotoSansSC"
    if font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"{assignment.title}-{student.name}",
        author="AhaMark",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=20,
        leading=26,
        alignment=TA_CENTER,
    )
    body = ParagraphStyle(
        "ChineseBody",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=9.5,
        leading=15,
        wordWrap="CJK",
    )
    heading = ParagraphStyle(
        "ChineseHeading", parent=body, fontSize=13, leading=20, spaceBefore=8, spaceAfter=5
    )
    story: list[Any] = [
        Paragraph("AhaMark 学生成绩报告", title),
        Spacer(1, 5 * mm),
        Table(
            [
                ["作业", assignment.title, "班级", school_class.name],
                ["学生", student.name, "学号", student.student_number],
                [
                    "总分",
                    f"{row.payload.total_score} / {row.payload.max_score}",
                    "发布版本",
                    f"v{release.version}",
                ],
                [
                    "生成时间",
                    now_utc().isoformat(timespec="seconds"),
                    "原卷图",
                    "本版本未嵌入原卷图，可在复核工作台查看",
                ],
            ],
            colWidths=[24 * mm, 58 * mm, 24 * mm, 63 * mm],
            style=[
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef2ff")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#eef2ff")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ],
        ),
        Paragraph("各题得分与教师反馈", heading),
    ]
    question_rows: list[list[Any]] = [["题号", "得分", "错误类型", "教师评语"]]
    knowledge: dict[str, list[Decimal]] = {}
    for detail in row.payload.details:
        question_rows.append(
            [
                Paragraph(detail.question_number, body),
                f"{detail.score} / {detail.max_score}",
                Paragraph(detail.final_error_type or "—", body),
                Paragraph(detail.final_feedback or "—", body),
            ]
        )
        rate = detail.score / detail.max_score
        for point_id in detail.knowledge_point_ids:
            knowledge.setdefault(str(point_id), []).append(rate)
    story.append(
        Table(
            question_rows,
            repeatRows=1,
            colWidths=[18 * mm, 26 * mm, 38 * mm, 87 * mm],
            style=[
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0e7ff")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ],
        )
    )
    story.extend([Paragraph("知识点表现", heading)])
    if knowledge:
        story.append(
            Table(
                [["知识点 ID", "平均得分率"]]
                + [
                    [key, f"{sum(values) / len(values) * 100:.1f}%"]
                    for key, values in knowledge.items()
                ],
                repeatRows=1,
                colWidths=[110 * mm, 45 * mm],
                style=[
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0e7ff")),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
                ],
            )
        )
    else:
        story.append(Paragraph("本作业没有可用的知识点关联数据。", body))
    story.extend(
        [
            PageBreak(),
            Paragraph("报告说明", heading),
            Paragraph(
                "本报告只读取发布版本固定的 complete ScoreSnapshot；"
                "后续成绩修改不会改变本报告对应的发布版本。"
                "主观题在没有真实评分 Provider 时由教师人工评分。",
                body,
            ),
        ]
    )
    document.build(story)
    return output.getvalue()


def create_analytics(db: Session, release: GradeRelease) -> AnalyticsSnapshot:
    scores = release_scores(db, release.id)
    knowledge_point_ids = {
        knowledge_point_id
        for row in scores
        for detail in row.payload.details
        for knowledge_point_id in detail.knowledge_point_ids
    }
    knowledge_point_names = {
        str(point.id): point.name
        for point in db.scalars(
            select(KnowledgePoint).where(
                KnowledgePoint.id.in_(knowledge_point_ids),
                KnowledgePoint.owner_id == release.owner_id,
            )
        )
    }
    snapshot = AnalyticsSnapshot(
        owner_id=release.owner_id,
        assignment_id=release.assignment_id,
        class_id=release.class_id,
        grade_release_id=release.id,
        source_snapshot_count=len(scores),
        schema_version="1.1",
        metrics=compute_metrics(scores, knowledge_point_names),
    )
    db.add(snapshot)
    db.flush()
    return snapshot
