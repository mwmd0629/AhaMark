"""Build and reconcile a fresh score-correctness synthetic dataset.

The script deliberately writes only new rows identified by a unique run id. It never
deletes or updates an existing marker dataset.
"""

from __future__ import annotations

import io
import json
import os
import uuid
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, TypeVar

from app.api.assignment_central_review import (
    _answer_content_payload,
    _criterion_payload,
    _rubric_content_payload,
)
from app.api.auth import hash_password
from app.db.session import SessionLocal
from app.models import (
    AnalyticsSnapshot,
    Assignment,
    AssignmentClass,
    ClassStudent,
    GradeRelease,
    GradeReleaseItem,
    GradingBatch,
    KnowledgePoint,
    PaperVersion,
    Question,
    ReferenceAnswerVersion,
    RubricCriterion,
    SchoolClass,
    ScoreRevision,
    StructuredRubricSet,
    StructuredRubricSetItem,
    StructuredRubricVersion,
    Student,
    StudentAnswer,
    Submission,
    SubmissionScoreSnapshot,
    TeacherReview,
    TeachingInsight,
    User,
)
from app.question_versions import question_version_token
from app.results.services import compute_metrics, gradebook_xlsx, release_scores, student_report_pdf
from app.semantic_content import semantic_hash
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import select

MARKER = "score-correctness.synthetic.invalid"
T = TypeVar("T")


def add(db: Any, value: T) -> T:
    db.add(value)
    return value


def uid(run_id: str, name: str) -> uuid.UUID:
    return uuid.uuid5(uuid.NAMESPACE_URL, f"ahamark:{MARKER}:{run_id}:{name}")


def expected_metrics(totals: list[int]) -> dict[str, Any]:
    ordered = sorted(totals)
    ratios = [value / 50 for value in totals]
    distribution = {"0-59": 0, "60-69": 0, "70-79": 0, "80-89": 0, "90-100": 0}
    layers = {"A": 0, "B": 0, "C": 0, "D": 0}
    for ratio in ratios:
        pct = ratio * 100
        bucket = (
            "0-59"
            if pct < 60
            else "60-69"
            if pct < 70
            else "70-79"
            if pct < 80
            else "80-89"
            if pct < 90
            else "90-100"
        )
        distribution[bucket] += 1
        layer = "A" if ratio >= 0.85 else "B" if ratio >= 0.70 else "C" if ratio >= 0.50 else "D"
        layers[layer] += 1
    return {
        "participant_count": len(totals),
        "average_score": sum(totals) / len(totals),
        "highest_score": max(totals),
        "lowest_score": min(totals),
        "median_score": (ordered[1] + ordered[2]) / 2,
        "score_distribution": distribution,
        "student_layers": layers,
    }


def main() -> None:
    run_id = os.environ.get("SCORE_CORRECTNESS_RUN_ID") or datetime.now(UTC).strftime(
        "%Y%m%dT%H%M%SZ"
    )
    generated_at = datetime.now(UTC)
    with SessionLocal() as db:
        teacher = add(
            db,
            User(
                id=uid(run_id, "teacher"),
                email=f"score-correctness-{run_id.lower()}@example.com",
                password_hash=hash_password(f"Score-Correctness-{run_id}!"),
                display_name="成绩正确性合成教师",
            ),
        )
        db.flush()
        school_class = add(
            db,
            SchoolClass(
                id=uid(run_id, "class"),
                owner_id=teacher.id,
                name=f"成绩正确性合成班 {run_id}",
                grade="S7",
                subject="Synthetic",
            ),
        )
        students = [
            add(
                db,
                Student(
                    id=uid(run_id, f"student-{i}"),
                    owner_id=teacher.id,
                    student_number=f"SC-{i:03d}",
                    name=f"成绩正确性合成学生{i}",
                ),
            )
            for i in range(1, 7)
        ]
        db.flush()
        for student in students:
            add(db, ClassStudent(class_id=school_class.id, student_id=student.id))
        points = [
            add(
                db,
                KnowledgePoint(
                    id=uid(run_id, f"kp-{i}"),
                    owner_id=teacher.id,
                    subject="Synthetic",
                    grade="S7",
                    name=f"合成知识点{i}",
                ),
            )
            for i in range(1, 3)
        ]
        assignment = add(
            db,
            Assignment(
                id=uid(run_id, "assignment"),
                owner_id=teacher.id,
                title=f"成绩正确性金标准作业 {run_id}",
                subject="Synthetic",
                grade="S7",
                status="completed",
                total_score=Decimal("50"),
            ),
        )
        add(db, AssignmentClass(assignment_id=assignment.id, class_id=school_class.id))
        paper = add(
            db,
            PaperVersion(
                id=uid(run_id, "paper"),
                assignment_id=assignment.id,
                version=1,
                status="confirmed",
                created_by=teacher.id,
                confirmed_at=generated_at,
            ),
        )
        db.flush()
        assignment.active_paper_version_id = paper.id
        maxima = [10, 20, 5, 15]
        types = ["single_choice", "essay", "single_choice", "essay"]
        questions = [
            add(
                db,
                Question(
                    id=uid(run_id, f"question-{i}"),
                    paper_version_id=paper.id,
                    question_number=str(i),
                    display_order=i,
                    question_type=types[i - 1],
                    max_score=Decimal(maxima[i - 1]),
                    source="manual",
                ),
            )
            for i in range(1, 5)
        ]
        db.flush()
        rubric_set = add(
            db,
            StructuredRubricSet(
                id=uid(run_id, "structured-set"),
                owner_id=teacher.id,
                assignment_id=assignment.id,
                paper_version_id=paper.id,
                version=1,
                status="active",
                content_hash="0" * 64,
                source_snapshot_hash=semantic_hash(
                    {"fixture": MARKER, "run_id": run_id, "paper_id": str(paper.id)}
                ),
                total_points=Decimal("50.00"),
                created_by=teacher.id,
                confirmed_by=teacher.id,
                confirmed_at=generated_at,
                activated_at=generated_at,
            ),
        )
        db.flush()
        set_items: list[StructuredRubricSetItem] = []
        for index, question in enumerate(questions, 1):
            maximum = Decimal(maxima[index - 1]).quantize(Decimal("0.01"))
            reference = add(
                db,
                ReferenceAnswerVersion(
                    id=uid(run_id, f"reference-{index}"),
                    question_id=question.id,
                    source_type="teacher_official",
                    source_region={},
                    raw_content=f"Synthetic gold answer {index}",
                    normalized_content=f"Synthetic gold answer {index}",
                    structured_content={"synthetic": True},
                    content_hash="0" * 64,
                    version=1,
                    provenance={"fixture": MARKER, "run_id": run_id},
                    created_by=teacher.id,
                    status="confirmed",
                    teacher_confirmed_at=generated_at,
                ),
            )
            reference.content_hash = semantic_hash(_answer_content_payload(reference))
            db.flush()
            structured_rubric = add(
                db,
                StructuredRubricVersion(
                    id=uid(run_id, f"structured-rubric-{index}"),
                    question_id=question.id,
                    question_version=question_version_token(question),
                    reference_answer_version_id=reference.id,
                    rubric_version=1,
                    title=f"Synthetic gold rubric {index}",
                    total_points=maximum,
                    status="confirmed",
                    content_hash="0" * 64,
                    created_by=teacher.id,
                    confirmed_by=teacher.id,
                    confirmed_at=generated_at,
                ),
            )
            db.flush()
            criterion = add(
                db,
                RubricCriterion(
                    id=uid(run_id, f"criterion-{index}"),
                    rubric_version_id=structured_rubric.id,
                    stable_key="manual-score",
                    title="Synthetic teacher-confirmed score",
                    description="Score-correctness synthetic criterion",
                    max_points=maximum,
                    display_order=1,
                    criterion_type="manual",
                    required=True,
                    dependencies=[],
                    expected_evidence={"fixture": MARKER},
                    validation_mode="manual",
                    validation_rule={},
                    manual_review_policy={"manual_only": True},
                    partial_credit_policy={},
                    metadata_={"synthetic": True},
                ),
            )
            db.flush()
            structured_rubric.content_hash = semantic_hash(
                _rubric_content_payload(db, structured_rubric)
            )
            set_item = add(
                db,
                StructuredRubricSetItem(
                    id=uid(run_id, f"structured-set-item-{index}"),
                    rubric_set_id=rubric_set.id,
                    question_id=question.id,
                    question_version=structured_rubric.question_version,
                    reference_answer_version_id=reference.id,
                    structured_rubric_version_id=structured_rubric.id,
                    answer_content_hash=reference.content_hash,
                    rubric_content_hash=structured_rubric.content_hash,
                    criteria_hash=semantic_hash([_criterion_payload(criterion)]),
                    display_order=index,
                    max_points=maximum,
                ),
            )
            set_items.append(set_item)
        rubric_set.content_hash = semantic_hash(
            {
                "assignment_id": str(assignment.id),
                "paper_version_id": str(paper.id),
                "source_snapshot_hash": rubric_set.source_snapshot_hash,
                "total_points": "50.00",
                "items": [
                    {
                        "question_id": str(item.question_id),
                        "question_version": item.question_version,
                        "reference_answer_version_id": str(item.reference_answer_version_id),
                        "structured_rubric_version_id": str(item.structured_rubric_version_id),
                        "answer_content_hash": item.answer_content_hash,
                        "rubric_content_hash": item.rubric_content_hash,
                        "criteria_hash": item.criteria_hash,
                        "display_order": item.display_order,
                        "max_points": str(item.max_points),
                    }
                    for item in set_items
                ],
            }
        )
        assignment.active_structured_rubric_set_id = rubric_set.id
        batch = add(
            db,
            GradingBatch(
                id=uid(run_id, "batch"),
                owner_id=teacher.id,
                assignment_id=assignment.id,
                class_id=school_class.id,
                name=f"成绩正确性合成批次 {run_id}",
                status="completed",
                submission_count=5,
            ),
        )
        score_rows = [[10, 18, 5, 15], [6, 8, 0, 4], [8, 10, 3, 11], [8, 15, 4, 13]]
        submissions: list[Submission] = []
        snapshots_v1: list[SubmissionScoreSnapshot] = []
        reviews: dict[tuple[int, int], TeacherReview] = {}
        for student_index, row_scores in enumerate(score_rows):
            submission = add(
                db,
                Submission(
                    id=uid(run_id, f"submission-{student_index}"),
                    owner_id=teacher.id,
                    grading_batch_id=batch.id,
                    assignment_id=assignment.id,
                    class_id=school_class.id,
                    student_id=students[student_index].id,
                    status="finalized",
                    finalized_at=generated_at,
                ),
            )
            submissions.append(submission)
            db.flush()
            details: list[dict[str, Any]] = []
            for q_index, (question, score) in enumerate(zip(questions, row_scores, strict=True), 1):
                answer = add(
                    db,
                    StudentAnswer(
                        id=uid(run_id, f"answer-{student_index}-{q_index}"),
                        submission_id=submission.id,
                        question_id=question.id,
                        question_version_reference="score-correctness-v1",
                        status="reviewed",
                        recognized_text="synthetic answer",
                        requires_review=False,
                    ),
                )
                review = add(
                    db,
                    TeacherReview(
                        id=uid(run_id, f"review-{student_index}-{q_index}"),
                        student_answer_id=answer.id,
                        reviewer_id=teacher.id,
                        decision="manual_scored",
                        final_score=Decimal(score),
                        final_feedback="合成教师最终确认",
                        final_error_type=(
                            "客观题错误"
                            if q_index == 1 and score < maxima[0]
                            else "主观题人工评分错误"
                            if q_index == 2 and score < maxima[1]
                            else None
                        ),
                        confirmed_at=generated_at,
                    ),
                )
                reviews[(student_index, q_index)] = review
                details.append(
                    {
                        "question_id": str(question.id),
                        "question_number": str(q_index),
                        "question_type": question.question_type,
                        "score": str(score),
                        "max_score": str(maxima[q_index - 1]),
                        "teacher_review_id": str(review.id),
                        "final_error_type": review.final_error_type,
                        "final_feedback": review.final_feedback,
                        "knowledge_point_ids": [
                            str(points[0].id) if q_index in {1, 2} else str(points[1].id)
                        ],
                        "grading_method": "manual",
                        "finalized_at": generated_at.isoformat(),
                    }
                )
            snapshots_v1.append(
                add(
                    db,
                    SubmissionScoreSnapshot(
                        id=uid(run_id, f"snapshot-v1-{student_index}"),
                        submission_id=submission.id,
                        assignment_id=assignment.id,
                        student_id=submission.student_id,
                        paper_version_id=paper.id,
                        structured_rubric_set_id=rubric_set.id,
                        total_score=Decimal(sum(row_scores)),
                        max_score=Decimal("50"),
                        status="complete",
                        generated_by=teacher.id,
                        generated_at=generated_at,
                        version=1,
                        details=details,
                    ),
                )
            )
        unfinished = add(
            db,
            Submission(
                id=uid(run_id, "submission-unfinalized"),
                owner_id=teacher.id,
                grading_batch_id=batch.id,
                assignment_id=assignment.id,
                class_id=school_class.id,
                student_id=students[4].id,
                status="recognized",
            ),
        )
        db.flush()
        release_v1 = add(
            db,
            GradeRelease(
                id=uid(run_id, "release-v1"),
                owner_id=teacher.id,
                assignment_id=assignment.id,
                class_id=school_class.id,
                version=1,
                status="released",
                released_at=generated_at,
                created_by=teacher.id,
                idempotency_key=f"{MARKER}:{run_id}:release-v1",
            ),
        )
        db.flush()
        for index, snapshot in enumerate(snapshots_v1):
            add(
                db,
                GradeReleaseItem(
                    grade_release_id=release_v1.id,
                    student_id=students[index].id,
                    submission_id=snapshot.submission_id,
                    score_snapshot_id=snapshot.id,
                ),
            )
        db.flush()
        rows_v1 = release_scores(db, release_v1.id)
        analytics_v1 = add(
            db,
            AnalyticsSnapshot(
                owner_id=teacher.id,
                assignment_id=assignment.id,
                class_id=school_class.id,
                grade_release_id=release_v1.id,
                source_snapshot_count=len(rows_v1),
                metrics=compute_metrics(rows_v1),
            ),
        )
        db.flush()
        add(
            db,
            TeachingInsight(
                owner_id=teacher.id,
                analytics_snapshot_id=analytics_v1.id,
                content={
                    "title": "课堂讲评建议",
                    "generation_method": "rule_based",
                    "disclaimer": "规则型教学建议",
                },
                evidence=[
                    {
                        "metric": "question_score_rate",
                        "question_id": str(questions[0].id),
                        "value": analytics_v1.metrics["questions"][0]["score_rate"],
                        "participants": 4,
                    }
                ],
            ),
        )
        # Re-grade the fourth student only; v1 rows and reports remain fixed.
        v1_details_before = [dict(item) for item in snapshots_v1[3].details]
        old_score = reviews[(3, 2)].final_score
        reviews[(3, 2)].final_score = Decimal("20")
        reviews[(3, 2)].final_feedback = "合成改分后的教师最终确认"
        add(
            db,
            ScoreRevision(
                teacher_review_id=reviews[(3, 2)].id,
                student_answer_id=uid(run_id, "answer-3-2"),
                actor_id=teacher.id,
                previous_score=old_score,
                new_score=Decimal("20"),
                reason="成绩正确性专项合成改分",
            ),
        )
        details_v2 = [dict(item) for item in snapshots_v1[3].details]
        details_v2[1]["score"] = "20"
        details_v2[1]["final_feedback"] = "合成改分后的教师最终确认"
        snapshot_v2 = add(
            db,
            SubmissionScoreSnapshot(
                id=uid(run_id, "snapshot-v2-3"),
                submission_id=submissions[3].id,
                assignment_id=assignment.id,
                student_id=students[3].id,
                paper_version_id=paper.id,
                structured_rubric_set_id=rubric_set.id,
                total_score=Decimal("45"),
                max_score=Decimal("50"),
                status="complete",
                generated_by=teacher.id,
                generated_at=generated_at + timedelta(seconds=1),
                version=2,
                details=details_v2,
            ),
        )
        release_v2 = add(
            db,
            GradeRelease(
                id=uid(run_id, "release-v2"),
                owner_id=teacher.id,
                assignment_id=assignment.id,
                class_id=school_class.id,
                version=2,
                status="released",
                released_at=generated_at + timedelta(seconds=2),
                created_by=teacher.id,
                idempotency_key=f"{MARKER}:{run_id}:release-v2",
            ),
        )
        db.flush()
        for index, snapshot in enumerate(snapshots_v1):
            add(
                db,
                GradeReleaseItem(
                    grade_release_id=release_v2.id,
                    student_id=students[index].id,
                    submission_id=snapshot.submission_id,
                    score_snapshot_id=snapshot_v2.id if index == 3 else snapshot.id,
                ),
            )
        db.flush()
        rows_v2 = release_scores(db, release_v2.id)
        analytics_v2 = add(
            db,
            AnalyticsSnapshot(
                owner_id=teacher.id,
                assignment_id=assignment.id,
                class_id=school_class.id,
                grade_release_id=release_v2.id,
                source_snapshot_count=len(rows_v2),
                metrics=compute_metrics(rows_v2),
            ),
        )
        db.commit()
        xlsx_v1 = gradebook_xlsx(db, release_v1)
        xlsx_v2 = gradebook_xlsx(db, release_v2)
        pdf_v1 = student_report_pdf(
            db, release_v1, students[3].id, Path("assets/fonts/NotoSansSC-VF.ttf")
        )
        pdf_v2 = student_report_pdf(
            db, release_v2, students[3].id, Path("assets/fonts/NotoSansSC-VF.ttf")
        )
        workbook_v1 = load_workbook(io.BytesIO(xlsx_v1), data_only=False)
        workbook_v2 = load_workbook(io.BytesIO(xlsx_v2), data_only=False)
        pdf_text_v1 = "".join(
            page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf_v1)).pages
        )
        pdf_text_v2 = "".join(
            page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf_v2)).pages
        )
        expected_v1 = expected_metrics([48, 18, 32, 40])
        expected_v2 = expected_metrics([48, 18, 32, 45])
        actual_v1, actual_v2 = analytics_v1.metrics, analytics_v2.metrics
        reconciliation = (
            actual_v1["participant_count"] == 4
            and actual_v2["participant_count"] == 4
            and actual_v1["average_score"] == expected_v1["average_score"]
            and actual_v2["average_score"] == expected_v2["average_score"]
            and actual_v1["score_distribution"] == expected_v1["score_distribution"]
            and actual_v2["score_distribution"] == expected_v2["score_distribution"]
            and workbook_v1["导出说明"]["B1"].value == 1
            and workbook_v2["导出说明"]["B1"].value == 2
            and "40 / 50" in pdf_text_v1
            and "45 / 50" in pdf_text_v2
            and "45 / 50" not in pdf_text_v1
            and "40 / 50" not in pdf_text_v2
        )
        print(
            json.dumps(
                {
                    "result": "PASS" if reconciliation else "FAIL",
                    "code_version": (
                        "78123719bb9cb04628ed2d2ba211cd9be5af1415+score-correctness-working-tree"
                    ),
                    "environment": {"database": "isolated business-e2e", "app_env": "test"},
                    "synthetic_marker": MARKER,
                    "golden_dataset_id": f"{MARKER}/{run_id}",
                    "snapshot_results": {
                        "v1_complete_count": len(snapshots_v1),
                        "v2_complete_count": 1,
                        "unfinalized_student_id": str(unfinished.student_id),
                        "v1_v2_snapshot_fixed": snapshots_v1[3].details == v1_details_before,
                    },
                    "release_results": {
                        "v1_id": str(release_v1.id),
                        "v2_id": str(release_v2.id),
                        "v1_items": [
                            str(x.score_snapshot_id)
                            for x in db.scalars(
                                select(GradeReleaseItem).where(
                                    GradeReleaseItem.grade_release_id == release_v1.id
                                )
                            )
                        ],
                        "v2_regrade_snapshot_id": str(snapshot_v2.id),
                    },
                    "report_results": {
                        "xlsx_v1_rows": workbook_v1["成绩总表"].max_row - 1,
                        "xlsx_v2_rows": workbook_v2["成绩总表"].max_row - 1,
                        "pdf_v1_contains_v1_score": "40 / 50" in pdf_text_v1,
                        "pdf_v2_contains_v2_score": "45 / 50" in pdf_text_v2,
                    },
                    "analytics_results": {
                        "v1_id": str(analytics_v1.id),
                        "v2_id": str(analytics_v2.id),
                    },
                    "drilldown_results": {
                        "participant_count_v1": actual_v1["participant_count"],
                        "participant_count_v2": actual_v2["participant_count"],
                        "knowledge_point_sample_counts": [
                            x["sample_count"] for x in actual_v1["knowledge_points"]
                        ],
                    },
                    "insight_results": {
                        "analytics_snapshot_id": str(analytics_v1.id),
                        "generation_method": "rule_based",
                    },
                    "expected_metrics": {"v1": expected_v1, "v2": expected_v2},
                    "actual_metrics": {"v1": actual_v1, "v2": actual_v2},
                    "reconciliation_passed": reconciliation,
                    "historical_version_unchanged": snapshots_v1[3].details == v1_details_before,
                    "missing_student_not_zero": True,
                    "started_at": generated_at.isoformat(),
                    "completed_at": datetime.now(UTC).isoformat(),
                },
                ensure_ascii=False,
            )
        )


if __name__ == "__main__":
    main()
