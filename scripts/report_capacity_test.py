"""Generate and verify reports for 50 distinct released synthetic students."""

import http.cookiejar
import io
import json
import statistics
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE = "http://127.0.0.1:8800"
MARKER = "performance-capacity.synthetic.invalid"
PASSWORD = "Synthetic-Capacity-Only!"


def uid(name: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_URL, f"ahamark:{MARKER}:{name}"))


def request(
    opener: urllib.request.OpenerDirector,
    method: str,
    path: str,
    csrf: str | None = None,
    payload: dict[str, Any] | None = None,
) -> tuple[dict[str, Any], float]:
    data = json.dumps(payload).encode() if payload is not None else None
    headers = {"Content-Type": "application/json"}
    if csrf:
        headers["X-CSRF-Token"] = csrf
    req = urllib.request.Request(BASE + path, data=data, headers=headers, method=method)
    started = time.perf_counter()
    try:
        with opener.open(req, timeout=30) as response:
            body = json.loads(response.read() or b"{}")
            body["_http_status"] = response.status
    except urllib.error.HTTPError as exc:
        body = json.loads(exc.read() or b"{}")
        body["_http_status"] = exc.code
    return body, (time.perf_counter() - started) * 1000


def milliseconds(start: str, end: str) -> float:
    return (datetime.fromisoformat(end) - datetime.fromisoformat(start)).total_seconds() * 1000


def download(opener: urllib.request.OpenerDirector, job_id: str) -> bytes:
    link, _ = request(opener, "GET", f"/api/report-jobs/{job_id}/download")
    with urllib.request.urlopen(link["url"], timeout=30) as response:
        return response.read()


def verify_xlsx(content: bytes) -> dict[str, Any]:
    sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
    rows = list(sheet.iter_rows(values_only=False))
    data_rows = rows[1:]
    student_numbers = [row[0] for row in data_rows]
    return {
        "rows": len(data_rows),
        "student_number_text_cells": sum(cell.data_type == "s" for cell in student_numbers),
        "unique_student_numbers": len({cell.value for cell in student_numbers}),
    }


def verify_zip(content: bytes) -> dict[str, Any]:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        names = archive.namelist()
        safe = all(
            name == Path(name).name and name.lower().endswith(".pdf") and ".." not in name
            for name in names
        )
        pdf_headers = sum(archive.read(name).startswith(b"%PDF") for name in names)
    return {
        "entries": len(names),
        "unique_names": len(set(names)),
        "safe_names": safe,
        "valid_pdf_headers": pdf_headers,
    }


def main() -> None:
    jar = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(jar))
    login, login_ms = request(
        opener,
        "POST",
        "/auth/login",
        payload={
            "email": f"capacity-teacher-1@{MARKER}",
            "password": PASSWORD,
        },
    )
    csrf = str(login["csrf_token"])
    release_id = uid("release-s1")
    run_id = str(int(time.time()))
    specs = [
        ("gradebook_xlsx", None),
        ("batch_student_reports", None),
        *[("student_report_pdf", uid(f"student-s1-{index}")) for index in range(1, 51)],
    ]
    created: list[dict[str, Any]] = []
    create_ms: list[float] = []
    started_all = time.perf_counter()
    for index, (report_type, student_id) in enumerate(specs):
        query = {
            "report_type": report_type,
            "idempotency_key": f"capacity-50-{run_id}-{index}",
        }
        if student_id:
            query["student_id"] = student_id
        job, elapsed = request(
            opener,
            "POST",
            f"/api/grade-releases/{release_id}/reports?{urllib.parse.urlencode(query)}",
            csrf,
        )
        created.append(job)
        create_ms.append(elapsed)
    deadline = time.monotonic() + 600
    terminal: dict[str, dict[str, Any]] = {}
    while len(terminal) < len(created) and time.monotonic() < deadline:
        for job in created:
            if job["id"] in terminal:
                continue
            current, _ = request(opener, "GET", f"/api/report-jobs/{job['id']}")
            if current["status"] in {
                "completed",
                "partially_completed",
                "failed",
                "expired",
            }:
                terminal[job["id"]] = current
        if len(terminal) < len(created):
            time.sleep(0.5)
    total_ms = (time.perf_counter() - started_all) * 1000
    completed = [job for job in terminal.values() if job["status"] == "completed"]
    pdf_jobs = [job for job in completed if job["report_type"] == "student_report_pdf"]
    xlsx_job = next(job for job in completed if job["report_type"] == "gradebook_xlsx")
    zip_job = next(job for job in completed if job["report_type"] == "batch_student_reports")
    pdf_durations = [milliseconds(job["started_at"], job["completed_at"]) for job in pdf_jobs]
    xlsx_content = download(opener, xlsx_job["id"])
    zip_content = download(opener, zip_job["id"])
    xlsx_validation = verify_xlsx(xlsx_content)
    zip_validation = verify_zip(zip_content)
    distinct_students = len({job["student_id"] for job in pdf_jobs})
    passed = (
        len(completed) == len(specs)
        and len(pdf_jobs) == 50
        and distinct_students == 50
        and xlsx_validation["rows"] == 50
        and xlsx_validation["student_number_text_cells"] == 50
        and zip_validation["entries"] == 50
        and zip_validation["unique_names"] == 50
        and zip_validation["safe_names"]
        and zip_validation["valid_pdf_headers"] == 50
    )
    evidence = {
        "schema_version": 2,
        "result": "passed" if passed else "failed",
        "scope": "50 distinct synthetic students in one fixed GradeRelease",
        "login_ms": round(login_ms, 2),
        "release_id": release_id,
        "report_jobs": {
            "created": len(created),
            "terminal": len(terminal),
            "completed": len(completed),
            "failed": len(terminal) - len(completed),
            "distinct_pdf_students": distinct_students,
            "create_api_p50_ms": round(statistics.median(create_ms), 2),
            "create_api_p95_ms": round(statistics.quantiles(create_ms, n=20)[18], 2),
            "pdf_execution_p50_ms": round(statistics.median(pdf_durations), 2),
            "pdf_execution_p95_ms": round(statistics.quantiles(pdf_durations, n=20)[18], 2),
            "total_completion_ms": round(total_ms, 2),
        },
        "xlsx": {
            **xlsx_validation,
            "size_bytes": len(xlsx_content),
            "execution_ms": round(
                milliseconds(xlsx_job["started_at"], xlsx_job["completed_at"]), 2
            ),
        },
        "zip": {
            **zip_validation,
            "size_bytes": len(zip_content),
            "execution_ms": round(milliseconds(zip_job["started_at"], zip_job["completed_at"]), 2),
        },
        "jobs": list(terminal.values()),
    }
    Path("docs/async-capacity-results.json").write_text(
        json.dumps(evidence, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps({key: value for key, value in evidence.items() if key != "jobs"}))


if __name__ == "__main__":
    main()
