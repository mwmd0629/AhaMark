import ast
import json
import uuid
from datetime import UTC, datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
from app.results.services import ValidatedScore, compute_metrics, gradebook_xlsx
from openpyxl import load_workbook

GOLDEN = json.loads(
    Path("tests/fixtures/score_correctness/golden.json").read_text(encoding="utf-8")
)


def test_score_correctness_seed_uses_structured_rubric_authority() -> None:
    path = Path("scripts/verify_score_correctness.py")
    source = path.read_text(encoding="utf-8")
    tree = ast.parse(source, filename=str(path))
    model_imports = {
        alias.name
        for node in ast.walk(tree)
        if isinstance(node, ast.ImportFrom) and node.module == "app.models"
        for alias in node.names
    }
    assert {"StructuredRubricSet", "StructuredRubricSetItem"} <= model_imports
    assert {"RubricVersion", "QuestionRubric", "RubricItem"}.isdisjoint(model_imports)
    for node in ast.walk(tree):
        if not isinstance(node, ast.Call) or not isinstance(node.func, ast.Name):
            continue
        if node.func.id == "SubmissionScoreSnapshot":
            fields = {keyword.arg for keyword in node.keywords}
            assert "structured_rubric_set_id" in fields
            assert "rubric_version_id" not in fields
    assert "active_structured_rubric_set_id" in source
    assert "active_rubric_version_id" not in source
    assert "question_version_token" in source
    assert "semantic_hash" in source


def _rows(version: str) -> list[ValidatedScore]:
    scores = GOLDEN[version]["student_totals"]
    question_scores = GOLDEN["v1"]["question_scores"]
    rows: list[ValidatedScore] = []
    for student_index, total in enumerate(scores):
        student_id = uuid.uuid5(uuid.NAMESPACE_URL, f"golden:{version}:{student_index}")
        details = []
        for number, maximum in GOLDEN["v1"]["question_max_scores"].items():
            raw = question_scores[number][student_index]
            if version == "v2" and student_index == 3 and number == "4":
                raw += 5
            question_id = uuid.uuid5(uuid.NAMESPACE_URL, f"golden:question:{number}")
            details.append(
                {
                    "question_id": question_id,
                    "question_number": number,
                    "question_type": "single_choice" if number in {"1", "3"} else "essay",
                    "score": raw,
                    "max_score": maximum,
                    "teacher_review_id": uuid.uuid5(
                        uuid.NAMESPACE_URL, f"golden:review:{version}:{student_index}:{number}"
                    ),
                    "final_error_type": "concept" if raw < maximum and number == "1" else None,
                    "final_feedback": "合成教师确认",
                    "knowledge_point_ids": [
                        uuid.uuid5(uuid.NAMESPACE_URL, f"golden:{key}")
                        for key, questions in GOLDEN["v1"]["knowledge_points"].items()
                        if number in questions
                    ],
                    "grading_method": "manual",
                    "finalized_at": datetime(2026, 7, 23, tzinfo=UTC),
                }
            )
        payload = SimpleNamespace(
            student_id=student_id,
            total_score=total,
            max_score=GOLDEN["assignment_max_score"],
            details=[SimpleNamespace(**detail) for detail in details],
        )
        rows.append(
            ValidatedScore(
                SimpleNamespace(
                    id=uuid.uuid5(uuid.NAMESPACE_URL, f"golden:snapshot:{version}:{student_index}"),
                    generated_at=datetime(2026, 7, 23, tzinfo=UTC),
                ),
                SimpleNamespace(id=uuid.uuid4()),
                payload,  # type: ignore[arg-type]
            )
        )
    return rows


def test_golden_metrics_are_independently_reconciled() -> None:
    for version in ("v1", "v2"):
        metrics = compute_metrics(_rows(version))
        expected = GOLDEN[version]["expected_metrics"]
        assert metrics["participant_count"] == expected["participant_count"]
        assert metrics["average_score"] == pytest.approx(expected["average_score"])
        assert metrics["highest_score"] == expected["highest_score"]
        assert metrics["lowest_score"] == expected["lowest_score"]
        assert metrics["median_score"] == pytest.approx(expected["median_score"])
        assert metrics["score_distribution"] == expected["score_distribution"]
        assert metrics["student_layers"] == expected["student_layers"]
        assert metrics["questions"][1]["correct_rate"] is None
        assert metrics["knowledge_points"][0]["sample_count"] == 4


def test_golden_xlsx_has_release_version_and_no_unfinished_students(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    rows = _rows("v1")
    monkeypatch.setattr("app.results.services.release_scores", lambda _db, _id: rows)
    names = {
        row.payload.student_id: SimpleNamespace(
            student_number=f"SC-{index:03d}", name=f"合成学生{index}"
        )
        for index, row in enumerate(rows, 1)
    }

    class FakeDb:
        def get(self, model: object, object_id: object) -> object:
            model_name = getattr(model, "__name__", "")
            if model_name == "Student":
                return names.get(object_id)
            if model_name == "SchoolClass":
                return SimpleNamespace(name="成绩正确性合成班")
            if model_name == "Assignment":
                return SimpleNamespace(title="成绩正确性金标准作业")
            if model_name == "KnowledgePoint":
                return SimpleNamespace(name=str(object_id))
            return None

    release = SimpleNamespace(
        id=uuid.uuid4(), class_id=uuid.uuid4(), assignment_id=uuid.uuid4(), version=1
    )
    workbook = load_workbook(
        __import__("io").BytesIO(gradebook_xlsx(FakeDb(), release)), data_only=False
    )
    sheet = workbook["成绩总表"]
    assert sheet.max_row == 5
    assert all(sheet.cell(row, 1).value != 0 for row in range(2, sheet.max_row + 1))
    assert workbook["导出说明"]["B1"].value == 1
