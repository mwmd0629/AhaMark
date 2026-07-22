import io
import uuid
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

import pytest
from app.results.services import (
    SnapshotPayload,
    ValidatedScore,
    compute_metrics,
    gradebook_xlsx,
    student_report_pdf,
)
from openpyxl import load_workbook
from pydantic import ValidationError
from pypdf import PdfReader


def payload(
    score: str = "8", maximum: str = "10", *, question_id: uuid.UUID | None = None
) -> SnapshotPayload:
    question_id = question_id or uuid.uuid4()
    return SnapshotPayload.model_validate(
        {
            "submission_id": uuid.uuid4(),
            "assignment_id": uuid.uuid4(),
            "student_id": uuid.uuid4(),
            "paper_version_id": uuid.uuid4(),
            "rubric_version_id": uuid.uuid4(),
            "total_score": score,
            "max_score": maximum,
            "question_count": 1,
            "details": [
                {
                    "question_id": question_id,
                    "question_number": "1",
                    "question_type": "single_choice",
                    "score": score,
                    "max_score": maximum,
                    "teacher_review_id": uuid.uuid4(),
                    "final_error_type": "concept" if score != maximum else None,
                    "knowledge_point_ids": [uuid.UUID(int=1)],
                    "grading_method": "manual",
                    "finalized_at": datetime.now(UTC),
                }
            ],
        }
    )


def validated(data: SnapshotPayload) -> ValidatedScore:
    snapshot = SimpleNamespace(id=uuid.uuid4(), generated_at=datetime.now(UTC))
    submission = SimpleNamespace(id=data.submission_id)
    return ValidatedScore(snapshot, submission, data)  # type: ignore[arg-type]


def test_snapshot_schema_rejects_score_out_of_range() -> None:
    with pytest.raises(ValidationError):
        payload("11", "10")


def test_snapshot_schema_rejects_duplicate_questions_and_total_mismatch() -> None:
    raw = payload().model_dump()
    raw["details"].append(raw["details"][0])
    raw["question_count"] = 2
    with pytest.raises(ValidationError, match="duplicate question_id"):
        SnapshotPayload.model_validate(raw)
    raw = payload().model_dump()
    raw["total_score"] = Decimal("7")
    with pytest.raises(ValidationError, match="total_score mismatch"):
        SnapshotPayload.model_validate(raw)


def test_metrics_formulas_distribution_errors_and_knowledge_points() -> None:
    question_id = uuid.uuid4()
    metrics = compute_metrics(
        [
            validated(payload("8", question_id=question_id)),
            validated(payload("10", question_id=question_id)),
        ]
    )
    assert metrics["participant_count"] == 2
    assert metrics["average_score"] == 9
    assert metrics["median_score"] == 9
    assert metrics["score_distribution"] == {
        "0-59": 0,
        "60-69": 0,
        "70-79": 0,
        "80-89": 1,
        "90-100": 1,
    }
    assert metrics["questions"][0]["score_rate"] == 0.9
    assert metrics["knowledge_points"][0]["mastery_rate"] == 0.9
    assert metrics["error_types"] == [{"code": "concept", "count": 1}]


def test_subjective_question_does_not_claim_correct_rate() -> None:
    data = payload()
    data.details[0].question_type = "essay"
    assert compute_metrics([validated(data)])["questions"][0]["correct_rate"] is None


def test_xlsx_has_required_sheets_and_student_number_is_text(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    data = payload()
    row = validated(data)
    monkeypatch.setattr("app.results.services.release_scores", lambda _db, _id: [row])
    student = SimpleNamespace(student_number="00123", name="测试学生")
    school_class = SimpleNamespace(name="一班")
    assignment = SimpleNamespace(title="测试作业")

    class FakeDb:
        def get(self, model: object, object_id: object) -> object:
            name = getattr(model, "__name__", "")
            return {"Student": student, "SchoolClass": school_class, "Assignment": assignment}.get(
                name
            )

    release = SimpleNamespace(
        id=uuid.uuid4(), class_id=uuid.uuid4(), assignment_id=uuid.uuid4(), version=2
    )
    content = gradebook_xlsx(FakeDb(), release)  # type: ignore[arg-type]
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    assert workbook.sheetnames == ["成绩总表", "题目统计", "知识点统计", "导出说明"]
    assert workbook["成绩总表"]["A2"].value == "00123"
    assert workbook["成绩总表"]["A2"].number_format == "@"
    assert workbook["导出说明"]["B1"].value == 2


def test_chinese_student_pdf_embeds_font_and_is_parseable(monkeypatch: pytest.MonkeyPatch) -> None:
    data = payload()
    data.details[0].final_feedback = "请复习一次函数概念。" * 80
    row = validated(data)
    monkeypatch.setattr("app.results.services.release_scores", lambda _db, _id: [row])
    student = SimpleNamespace(student_number="00123", name="合成测试学生")
    school_class = SimpleNamespace(name="八年级一班")
    assignment = SimpleNamespace(title="函数综合练习")

    class FakeDb:
        def get(self, model: object, object_id: object) -> object:
            return {
                "Student": student,
                "SchoolClass": school_class,
                "Assignment": assignment,
            }.get(getattr(model, "__name__", ""))

    release = SimpleNamespace(
        id=uuid.uuid4(), class_id=uuid.uuid4(), assignment_id=uuid.uuid4(), version=3
    )
    font = Path("apps/api/assets/fonts/NotoSansSC-VF.ttf")
    content = student_report_pdf(FakeDb(), release, data.student_id, font)  # type: ignore[arg-type]
    reader = PdfReader(io.BytesIO(content))
    text = "".join(page.extract_text() or "" for page in reader.pages)
    assert len(reader.pages) >= 2
    assert "AhaMark" in text and "合成测试学生" in text and "函数综合练习" in text
    assert b"NotoSansSC" in content
